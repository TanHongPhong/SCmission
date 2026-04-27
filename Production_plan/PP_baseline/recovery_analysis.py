"""
Multi-Level Recovery Analysis
==============================
For each recovery level (0%, 70%, 80%, 90%, 95%, 100%):
  - "Recovery" = how close we get back to baseline shortage
  - Minimize AIR cost subject to shortage <= target
  - Output: comparison table + Excel

Recovery formula:
  worst_shortage = shortage with NO mitigation (only reflash from May 6)
  baseline_shortage = shortage without scenario (from audit_current_plan.py)
  gap = worst_shortage - baseline_shortage
  target = worst_shortage - recovery% x gap
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pulp import (
    LpProblem, LpMinimize, LpVariable, lpSum, LpInteger,
    PULP_CBC_CMD, LpStatus, value
)
from datetime import datetime, date, timedelta
from collections import defaultdict
import pandas as pd

INPUT_FILE  = "SCM_round2.1_new.xlsx"
OUTPUT_FILE = "SCM_round2.1_recovery_analysis.xlsx"
PLAN_SHEET  = "Production_plan"
INPUT_SHEET = "input"

# Scenario constants
PCBA_SKU     = "A"
PCBA_BOM     = 2
PCBA_ON_HOLD = 370_000
REFLASH_WAIT = 21
SUPPLIER_CAP = 4_000
AIR_LEAD     = 3
SEA_LEAD     = 21
AIR_COST     = 2.0

RECOVERY_LEVELS = [0, 50, 70, 80, 90, 95, 100]


def to_date(x):
    if isinstance(x, datetime): return x.date()
    if isinstance(x, date):     return x
    return None

def num(x, default=0):
    if x is None or str(x).strip() == "": return default
    try: return float(x)
    except: return default

def vi(v):
    return int(round(v)) if v is not None else 0


def read_data():
    """Read all data once."""
    # Priorities
    df = pd.read_excel(INPUT_FILE, sheet_name=INPUT_SHEET, header=None)
    priority_map = {}
    for r in range(df.shape[0]):
        sku, pri = df.iloc[r, 0], df.iloc[r, 6]
        if pd.notna(sku) and pd.notna(pri) and str(sku).strip() not in ('SKU', ''):
            try: priority_map[str(sku).strip()] = int(pri)
            except: pass

    # Production plan
    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb[PLAN_SHEET]

    date_cols = []
    col = 11
    while True:
        try: d = to_date(ws.cell(42, col).value)
        except: d = None
        if d is None: break
        date_cols.append((col, d))
        col += 1

    rows = []
    r = 43
    while True:
        sku = ws.cell(r, 1).value
        if sku is None or str(sku).strip() == "": break
        rows.append({
            "row": r,
            "sku": str(sku).strip(),
            "fixture": str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else "N/A",
            "weekly_cap": float(num(ws.cell(r, 6).value, 0)),
            "ship_request": to_date(ws.cell(r, 7).value),
            "demand": int(num(ws.cell(r, 8).value, 0)),
        })
        r += 1

    sku_groups = defaultdict(list)
    sku_fixture = {}
    for rd in rows:
        sku_groups[rd["sku"]].append(rd)
        sku_fixture[rd["sku"]] = rd["fixture"]
    for s in sku_groups:
        sku_groups[s].sort(key=lambda x: x["ship_request"])
    all_skus = sorted(sku_groups.keys())

    fixture_max_cap = {}
    for rd in rows:
        fx = rd["fixture"]
        if fx != "N/A":
            if fx not in fixture_max_cap or rd["weekly_cap"] > fixture_max_cap[fx]:
                fixture_max_cap[fx] = rd["weekly_cap"]
    fixture_daily_cap = {fx: cap / 6.0 for fx, cap in fixture_max_cap.items()}

    return rows, sku_groups, all_skus, sku_fixture, fixture_daily_cap, date_cols, priority_map


def solve_scenario(rows, sku_groups, all_skus, sku_fixture, fixture_daily_cap,
                   date_cols, priority_map,
                   max_shortage_a=None, no_new_supply=False):
    """
    Solve production plan with PCBA scenario.
    - max_shortage_a: if set, constrain SKU A shortage <= this value
    - no_new_supply: if True, force air=0 and sea=0 (only reflash)
    Returns dict of metrics.
    """
    num_days = len(date_cols)
    horizon_start = date_cols[0][1]
    max_cal = (date_cols[-1][1] - horizon_start).days

    model = LpProblem("Scenario", LpMinimize)

    # Variables
    prod = {(s, t): LpVariable(f"p_{s}_{t}", 0, cat=LpInteger)
            for s in all_skus for t in range(num_days)}
    inv   = {rd["row"]: LpVariable(f"inv_{rd['row']}",   0, cat=LpInteger) for rd in rows}
    short = {rd["row"]: LpVariable(f"short_{rd['row']}", 0, cat=LpInteger) for rd in rows}

    air_ship = {d: LpVariable(f"air_{d}", 0, cat=LpInteger) for d in range(max_cal + 1)}
    sea_ship = {d: LpVariable(f"sea_{d}", 0, cat=LpInteger) for d in range(max_cal + 1)}

    for d in range(max_cal + 1):
        if no_new_supply:
            model += air_ship[d] == 0
            model += sea_ship[d] == 0
        else:
            model += air_ship[d] + sea_ship[d] <= SUPPLIER_CAP

    # Objective: minimize air cost + tiny inventory/shortage penalty for tie-breaking
    # We want MINIMUM AIR COST for a given service level
    obj = [AIR_COST * lpSum(air_ship[d] for d in range(max_cal + 1))]
    for rd in rows:
        pri = priority_map.get(rd["sku"], 2)
        pen = 10 if pri == 1 else 1
        obj.append(0.1 * pen * short[rd["row"]]) # Low weight so we focus on AIR cost
    obj.append(0.0001 * lpSum(inv[rd["row"]] for rd in rows))
    model += lpSum(obj)

    # Fixture capacity
    for fx, dc in fixture_daily_cap.items():
        if fx == "N/A" or dc <= 0: continue
        sfx = [s for s in all_skus if sku_fixture.get(s) == fx]
        for t in range(num_days):
            model += lpSum(prod[(s, t)] for s in sfx) <= dc

    # Inventory balance
    for s, periods in sku_groups.items():
        for j, rd in enumerate(periods):
            i = rd["row"]; ship = rd["ship_request"]
            if j == 0:
                beg = 0
                pp = lpSum(prod[(s, t)] for t, (_, d) in enumerate(date_cols) if d <= ship)
            else:
                beg = inv[periods[j - 1]["row"]]
                ps = periods[j - 1]["ship_request"]
                pp = lpSum(prod[(s, t)] for t, (_, d) in enumerate(date_cols)
                           if d > ps and d <= ship)
            model += inv[i] - short[i] == beg + pp - rd["demand"]

    # PCBA supply constraint
    for t, (_, dt) in enumerate(date_cols):
        cal = (dt - horizon_start).days
        reflash = PCBA_ON_HOLD if cal >= REFLASH_WAIT else 0
        cumul_new = (
            lpSum(air_ship[d] for d in range(max_cal + 1) if d + AIR_LEAD <= cal) +
            lpSum(sea_ship[d] for d in range(max_cal + 1) if d + SEA_LEAD <= cal)
        )
        cumul_a = lpSum(prod[(PCBA_SKU, t2)] for t2 in range(t + 1))
        model += cumul_a * PCBA_BOM <= reflash + cumul_new

    # Max shortage constraint for SKU A specifically
    if max_shortage_a is not None:
        model += lpSum(short[rd["row"]] for rd in rows if rd["sku"] == PCBA_SKU) <= max_shortage_a

    PULP_CBC_CMD(msg=False, timeLimit=120).solve(model)

    # Extract
    tot_short = sum(vi(value(short[rd["row"]])) for rd in rows)
    tot_plan  = sum(vi(value(prod[(s, t)])) for s in all_skus for t in range(num_days))
    t_air     = sum(vi(value(air_ship[d])) for d in range(max_cal + 1))
    t_sea     = sum(vi(value(sea_ship[d])) for d in range(max_cal + 1))

    sku_short = {}
    sku_plan  = {}
    for s in all_skus:
        sku_short[s] = sum(vi(value(short[rd["row"]])) for rd in rows if rd["sku"] == s)
        sku_plan[s]  = sum(vi(value(prod[(s, t)])) for t in range(num_days))

    return {
        "status":     LpStatus[model.status],
        "tot_plan":   tot_plan,
        "tot_short":  tot_short,
        "air_pcba":   t_air,
        "sea_pcba":   t_sea,
        "air_cost":   t_air * AIR_COST,
        "sku_short":  sku_short,
        "sku_plan":   sku_plan,
    }


def solve_baseline_a(rows, sku_groups, all_skus, sku_fixture, fixture_daily_cap,
                     date_cols, priority_map):
    """Solve with HIGH shortage penalty (like main.py 100% recovery) to find
    the minimum possible A shortage = best case."""
    num_days = len(date_cols)
    horizon_start = date_cols[0][1]
    max_cal = (date_cols[-1][1] - horizon_start).days

    model = LpProblem("Baseline", LpMinimize)
    prod = {(s, t): LpVariable("p_{}_{}".format(s, t), 0, cat=LpInteger)
            for s in all_skus for t in range(num_days)}
    inv   = {rd["row"]: LpVariable("inv_{}".format(rd['row']),   0, cat=LpInteger) for rd in rows}
    short = {rd["row"]: LpVariable("short_{}".format(rd['row']), 0, cat=LpInteger) for rd in rows}

    air_ship = {d: LpVariable("air_{}".format(d), 0, cat=LpInteger) for d in range(max_cal + 1)}
    sea_ship = {d: LpVariable("sea_{}".format(d), 0, cat=LpInteger) for d in range(max_cal + 1)}
    for d in range(max_cal + 1):
        model += air_ship[d] + sea_ship[d] <= SUPPLIER_CAP

    # HIGH shortage penalty -> solver minimizes shortage first, then cost
    obj = []
    for rd in rows:
        pri = priority_map.get(rd["sku"], 2)
        pen = 1_000_000 if pri == 1 else 10_000
        obj.append(pen * short[rd["row"]])
    obj.append(0.001 * lpSum(inv[rd["row"]] for rd in rows))
    obj.append(AIR_COST * lpSum(air_ship[d] for d in range(max_cal + 1)))
    model += lpSum(obj)

    for fx, dc in fixture_daily_cap.items():
        if fx == "N/A" or dc <= 0: continue
        sfx = [s for s in all_skus if sku_fixture.get(s) == fx]
        for t in range(num_days):
            model += lpSum(prod[(s, t)] for s in sfx) <= dc

    for s, periods in sku_groups.items():
        for j, rd in enumerate(periods):
            i = rd["row"]; ship = rd["ship_request"]
            if j == 0:
                beg = 0
                pp = lpSum(prod[(s, t)] for t, (_, d) in enumerate(date_cols) if d <= ship)
            else:
                beg = inv[periods[j - 1]["row"]]
                ps = periods[j - 1]["ship_request"]
                pp = lpSum(prod[(s, t)] for t, (_, d) in enumerate(date_cols)
                           if d > ps and d <= ship)
            model += inv[i] - short[i] == beg + pp - rd["demand"]

    for t, (_, dt) in enumerate(date_cols):
        cal = (dt - horizon_start).days
        reflash = PCBA_ON_HOLD if cal >= REFLASH_WAIT else 0
        cumul_new = (
            lpSum(air_ship[d] for d in range(max_cal + 1) if d + AIR_LEAD <= cal) +
            lpSum(sea_ship[d] for d in range(max_cal + 1) if d + SEA_LEAD <= cal))
        cumul_a = lpSum(prod[(PCBA_SKU, t2)] for t2 in range(t + 1))
        model += cumul_a * PCBA_BOM <= reflash + cumul_new

    PULP_CBC_CMD(msg=False, timeLimit=120).solve(model)

    tot_short = sum(vi(value(short[rd["row"]])) for rd in rows)
    tot_plan  = sum(vi(value(prod[(s, t)])) for s in all_skus for t in range(num_days))
    t_air     = sum(vi(value(air_ship[d])) for d in range(max_cal + 1))
    t_sea     = sum(vi(value(sea_ship[d])) for d in range(max_cal + 1))
    sku_short = {}
    sku_plan  = {}
    for s in all_skus:
        sku_short[s] = sum(vi(value(short[rd["row"]])) for rd in rows if rd["sku"] == s)
        sku_plan[s]  = sum(vi(value(prod[(s, t)])) for t in range(num_days))
    return {
        "status": LpStatus[model.status], "tot_plan": tot_plan,
        "tot_short": tot_short, "air_pcba": t_air, "sea_pcba": t_sea,
        "air_cost": t_air * AIR_COST, "sku_short": sku_short, "sku_plan": sku_plan,
    }


def main():
    print("=" * 70)
    print("  MULTI-LEVEL RECOVERY ANALYSIS (SKU A FOCUS)")
    print("=" * 70)

    data = read_data()
    rows      = data[0]
    all_skus  = data[2]
    num_days  = len(data[5])
    total_dem = sum(rd["demand"] for rd in rows)

    # ── Step 1: Worst case (no air/sea, reflash only) ───────
    print("\n[1/3] Solving WORST CASE (no air, no sea, reflash only) ...")
    worst = solve_scenario(*data, no_new_supply=True)
    worst_a = worst['sku_short'].get('A', 0)
    print("       A shortage = {:,}  (total = {:,})".format(worst_a, worst['tot_short']))

    # ── Step 2: Best case (minimize A shortage, then cost) ──
    print("[2/3] Solving BEST CASE (minimize A shortage with air) ...")
    best = solve_baseline_a(*data)
    best_a = best['sku_short'].get('A', 0)
    print("       A shortage = {:,}  (total = {:,})".format(best_a, best['tot_short']))
    print("       Air cost   = ${:,.2f}  ({:,} PCBA)".format(best['air_cost'], best['air_pcba']))

    gap_a = worst_a - best_a
    print("\n       Gap A = {:,} - {:,} = {:,} FG units".format(worst_a, best_a, gap_a))

    if gap_a <= 0:
        print("\n  [ERROR] No gap to recover. Air cannot help A.")
        return

    # ── Step 3: Each recovery level ─────────────────────────
    levels = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 95, 100]
    print("\n[3/3] Solving {} recovery levels ...\n".format(len(levels)))

    results = []
    for pct in levels:
        target_a = worst_a - int(pct / 100.0 * gap_a)
        if pct == 0:
            res = worst.copy()
            res["target_a"] = target_a
        elif pct == 100:
            res = best.copy()
            res["target_a"] = target_a
        else:
            # Minimize air cost subject to A shortage <= target_a
            res = solve_scenario(*data, max_shortage_a=target_a)
            res["target_a"] = target_a
        res["pct"] = pct
        results.append(res)
        a_act = res['sku_short'].get('A', 0)
        print("  {:>3}%: A target={:>7,}  A actual={:>7,}  "
              "air={:>6,}  cost=${:>10,.2f}  status={}".format(
            pct, target_a, a_act, res['air_pcba'], res['air_cost'], res['status']))



    # ── Output table ────────────────────────────────────────
    a_dem = sum(rd["demand"] for rd in rows if rd["sku"] == "A")
    print()
    print("=" * 115)
    print("  SKU A RECOVERY LEVEL COMPARISON")
    print("=" * 115)
    hdr = "{:>8} {:>9} {:>9} {:>9} {:>9} {:>8} {:>8} {:>12} {:>8} {:>8}"
    print(hdr.format("Recov%", "TargetA", "ActualA", "A Short%", "TotShort",
                     "Air PCBA", "Sea PCBA", "Air Cost",
                     "A Plan", "B Short"))
    print("-" * 115)
    for r in results:
        a_sh = r['sku_short'].get('A', 0)
        print(hdr.format(
            "{}%".format(r['pct']),
            "{:,}".format(r['target_a']),
            "{:,}".format(a_sh),
            "{:.1f}%".format(100 * a_sh / a_dem) if a_dem else "0%",
            "{:,}".format(r['tot_short']),
            "{:,}".format(r['air_pcba']),
            "{:,}".format(r['sea_pcba']),
            "${:,.0f}".format(r['air_cost']),
            "{:,}".format(r['sku_plan'].get('A', 0)),
            "{:,}".format(r['sku_short'].get('B', 0)),
        ))
    print("-" * 115)

    # Cost per unit recovered
    if gap_a > 0 and best['air_cost'] > 0:
        print("\n  Cost efficiency:")
        print("    Full recovery cost : ${:,.2f} for {:,} units recovered".format(
            best['air_cost'], gap_a))
        print("    Avg cost per unit  : ${:.2f} / FG recovered".format(
            best['air_cost'] / gap_a))

    # ── Save to Excel ───────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Recovery_Analysis"

    orange = PatternFill("solid", fgColor="E65100")
    green  = PatternFill("solid", fgColor="E2F0D9")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red    = PatternFill("solid", fgColor="FDE9D9")
    wh_b   = Font(color="FFFFFF", bold=True, size=11)
    bld    = Font(bold=True)
    thin   = Side(style="thin", color="BFBFBF")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Recovery %", "A Target Short", "A Actual Short", "A Short %",
        "Total Shortage", "Total Planned",
        "Air PCBA", "Sea PCBA", "Air Cost (USD)",
    ]
    for s in all_skus:
        headers.append("{} Planned".format(s))
        headers.append("{} Shortage".format(s))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font = orange, wh_b
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bdr

    for ri, r in enumerate(results, 2):
        a_sh = r['sku_short'].get('A', 0)
        vals = [
            "{}%".format(r['pct']),
            r["target_a"],
            a_sh,
            "{:.1f}%".format(100 * a_sh / a_dem) if a_dem else "0%",
            r["tot_short"],
            r["tot_plan"],
            r["air_pcba"],
            r["sea_pcba"],
            r["air_cost"],
        ]
        for s in all_skus:
            vals.append(r["sku_plan"].get(s, 0))
            vals.append(r["sku_short"].get(s, 0))

        for ci, v in enumerate(vals, 1):
            cl = ws.cell(ri, ci)
            cl.value = v
            cl.border = bdr
        if r["pct"] >= 95:
            ws.cell(ri, 1).fill = green
        elif r["pct"] >= 80:
            ws.cell(ri, 1).fill = yellow
        else:
            ws.cell(ri, 1).fill = red

    for col_cells in ws.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[cl].width = min(max(mx + 3, 12), 22)

    # ── Summary sheet ───────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    summary = [
        ["SKU A Recovery Analysis", ""],
        ["", ""],
        ["Total Demand (all SKUs)", "{:,}".format(total_dem)],
        ["SKU A Demand", "{:,}".format(a_dem)],
        ["Planning Horizon", "{} workdays".format(num_days)],
        ["", ""],
        ["Scenario Parameters", ""],
        ["PCBA on hold (defective)", "{:,}".format(PCBA_ON_HOLD)],
        ["Reflash wait", "{} days (calendar)".format(REFLASH_WAIT)],
        ["Supplier capacity", "{:,} PCBA/day".format(SUPPLIER_CAP)],
        ["Air lead time", "{} days".format(AIR_LEAD)],
        ["Sea lead time", "{} days".format(SEA_LEAD)],
        ["Air cost premium", "${}/PCBA".format(AIR_COST)],
        ["BOM usage (PCBA per FG)", "{}".format(PCBA_BOM)],
        ["", ""],
        ["Reference Points (SKU A)", ""],
        ["Worst case A shortage (0% recovery)", "{:,}".format(worst_a)],
        ["Best case A shortage (100% recovery)", "{:,}".format(best_a)],
        ["Gap to recover", "{:,} FG units".format(gap_a)],
        ["Full recovery air cost", "${:,.2f}".format(best['air_cost'])],
        ["Avg cost per unit recovered", "${:.2f} / FG".format(best['air_cost'] / gap_a) if gap_a else "N/A"],
    ]
    for ri, (k, v) in enumerate(summary, 1):
        ws2.cell(ri, 1).value = k
        ws2.cell(ri, 2).value = v
        ws2.cell(ri, 1).border = bdr
        ws2.cell(ri, 2).border = bdr
        if v == "" and k: ws2.cell(ri, 1).font = bld
    ws2.cell(1, 1).fill, ws2.cell(1, 1).font = orange, wh_b
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 30

    wb.save(OUTPUT_FILE)
    print("\n  Output saved to: {}".format(OUTPUT_FILE))
    print("=" * 115)


if __name__ == "__main__":
    main()
