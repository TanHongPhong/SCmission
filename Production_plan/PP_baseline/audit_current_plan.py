"""
Baseline Production Plan (BEFORE scenario disruption)
=====================================================
Production plan under normal operating conditions:
  - Fixture capacity constraints (daily = weekly / 6)
  - SKU priority allocation (A>B, C>D, E>F from input sheet)
  - Inventory continuity across demand periods
  - No earliest-date limit (production can start from day 1)
  - No material/BOM constraints (assumed sufficient supply)

This represents the ORIGINAL plan before any scenario pop-up.
Compare with main.py which handles the scenario (e.g. firmware defect).
"""

import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pulp import (
    LpProblem, LpMinimize, LpVariable, lpSum, LpInteger,
    PULP_CBC_CMD, LpStatus, value
)
from datetime import datetime, date, timedelta
from collections import defaultdict
import pandas as pd

INPUT_FILE = "SCM_round2.1_new.xlsx"
OUTPUT_FILE = "SCM_round2.1_baseline_solved.xlsx"
PLAN_SHEET = "Production_plan"
INPUT_SHEET = "input"

# Priority-dependent shortage penalties
# 100× gap ensures solver NEVER lets pri-1 have shortage
# if pri-2 could absorb it on the same fixture
M1_PRI1 = 1_000_000   # Priority-1 shortage penalty  (very high)
M1_PRI2 = 10_000       # Priority-2 shortage penalty  (high)
M2      = 0.001        # Inventory holding penalty     (tiny, for JIT preference)

INITIAL_FG_INV = 0


# ── Helpers ─────────────────────────────────────────────────────

def to_date(x):
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    return None


def num(x, default=0):
    if x is None or str(x).strip() == "":
        return default
    try:
        return float(x)
    except:
        return default


def solve_as_int(v):
    if v is None:
        return 0
    return int(round(v))


# ── Read priorities from input sheet ────────────────────────────

def read_priorities(file_path, sheet_name):
    """Read SKU priority from the 'input' sheet."""
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    priority_map = {}
    for r in range(df.shape[0]):
        sku = df.iloc[r, 0]
        pri = df.iloc[r, 6]
        if pd.notna(sku) and pd.notna(pri) and str(sku).strip() not in ('SKU', ''):
            try:
                priority_map[str(sku).strip()] = int(pri)
            except (ValueError, TypeError):
                pass
    return priority_map


# ── Main solver ─────────────────────────────────────────────────

def main():
    # ── 1. Read priorities ──────────────────────────────────────
    priority_map = read_priorities(INPUT_FILE, INPUT_SHEET)
    print("[INFO] SKU priorities:")
    for s, p in sorted(priority_map.items()):
        print(f"       {s}: priority {p}")

    # ── 2. Read production plan data ────────────────────────────
    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb[PLAN_SHEET]

    date_cols = []
    col = 11
    while True:
        try:
            d = to_date(ws.cell(42, col).value)
        except Exception:
            d = None
        if d is None:
            break
        date_cols.append((col, d))
        col += 1
    num_days = len(date_cols)
    print(f"[INFO] {num_days} workdays  ({date_cols[0][1]} to {date_cols[-1][1]})")

    rows = []
    r = 43
    while True:
        sku = ws.cell(r, 1).value
        if sku is None or str(sku).strip() == "":
            break
        rows.append({
            "row": r,
            "sku": str(sku).strip(),
            "description": str(ws.cell(r, 2).value or ""),
            "market": str(ws.cell(r, 3).value or ""),
            "product_group": str(ws.cell(r, 4).value or ""),
            "fixture": (str(ws.cell(r, 5).value).strip()
                        if ws.cell(r, 5).value else "N/A"),
            "weekly_cap": float(num(ws.cell(r, 6).value, 0)),
            "ship_request": to_date(ws.cell(r, 7).value),
            "demand": int(num(ws.cell(r, 8).value, 0)),
        })
        r += 1
    print(f"[INFO] {len(rows)} demand lines")

    # ── 3. Group & capacities ───────────────────────────────────
    sku_groups = defaultdict(list)
    sku_fixture = {}
    for rd in rows:
        sku_groups[rd["sku"]].append(rd)
        sku_fixture[rd["sku"]] = rd["fixture"]
    for s in sku_groups:
        sku_groups[s].sort(key=lambda x: x["ship_request"])

    all_skus = sorted(sku_groups.keys())

    fixture_max_cap = {}
    for rd in rows:
        fx = rd["fixture"]
        if fx != "N/A":
            if fx not in fixture_max_cap or rd["weekly_cap"] > fixture_max_cap[fx]:
                fixture_max_cap[fx] = rd["weekly_cap"]
    fixture_daily_cap = {fx: cap / 6.0 for fx, cap in fixture_max_cap.items()}

    print("[INFO] Fixture daily caps:")
    for fx, cap in sorted(fixture_daily_cap.items()):
        skus_on = [s for s in all_skus if sku_fixture.get(s) == fx]
        print(f"       {fx}: {cap:.1f} /day  "
              f"(SKUs: {', '.join(s + '(P' + str(priority_map.get(s, '?')) + ')' for s in skus_on)})")

    # ── 4. Build LP with priority-weighted shortage ─────────────
    model = LpProblem("Scenario_Production_Plan", LpMinimize)

    # Decision: production per SKU per day
    prod = {}
    for s in all_skus:
        for t in range(num_days):
            prod[(s, t)] = LpVariable(f"p_{s}_d{t}", 0, cat=LpInteger)

    # State: ending inventory & shortage per demand-period
    inv   = {}
    short = {}
    for rd in rows:
        i = rd["row"]
        inv[i]   = LpVariable(f"inv_r{i}",   0, cat=LpInteger)
        short[i] = LpVariable(f"short_r{i}", 0, cat=LpInteger)

    # ── Objective: priority-weighted shortage + tiny inv hold ───
    # Priority-1 shortage penalty >> Priority-2
    # → solver will always protect pri-1 SKUs first
    obj_terms = []
    for rd in rows:
        i = rd["row"]
        pri = priority_map.get(rd["sku"], 2)
        penalty = M1_PRI1 if pri == 1 else M1_PRI2
        obj_terms.append(penalty * short[i])
    obj_terms.append(M2 * lpSum(inv[rd["row"]] for rd in rows))
    model += lpSum(obj_terms)

    # ── Constraint A: fixture daily capacity ────────────────────
    for fx, daily_cap in fixture_daily_cap.items():
        if fx == "N/A" or daily_cap <= 0:
            continue
        skus_fx = [s for s in all_skus if sku_fixture.get(s) == fx]
        for t in range(num_days):
            model += lpSum(prod[(s, t)] for s in skus_fx) <= daily_cap

    # ── Constraint B: inventory balance per SKU per period ──────
    for s, periods in sku_groups.items():
        for j, rd in enumerate(periods):
            i    = rd["row"]
            ship = rd["ship_request"]

            if j == 0:
                beg = INITIAL_FG_INV
                period_prod = lpSum(
                    prod[(s, t)]
                    for t, (_, d) in enumerate(date_cols) if d <= ship
                )
            else:
                prev_ship = periods[j - 1]["ship_request"]
                beg       = inv[periods[j - 1]["row"]]
                period_prod = lpSum(
                    prod[(s, t)]
                    for t, (_, d) in enumerate(date_cols)
                    if d > prev_ship and d <= ship
                )

            model += inv[i] - short[i] == beg + period_prod - rd["demand"]

    # ── 5. Solve ────────────────────────────────────────────────
    print("[INFO] Solving with priority weights …")
    solver = PULP_CBC_CMD(msg=False, timeLimit=120)
    model.solve(solver)
    print(f"[INFO] Status: {LpStatus[model.status]}")

    # ── 6. Build output workbook ────────────────────────────────
    wb_out  = Workbook()
    ws_out  = wb_out.active
    ws_out.title = "Scenario_Result"

    # Styling
    navy   = PatternFill("solid", fgColor="1F4E78")
    teal   = PatternFill("solid", fgColor="0F766E")
    green  = PatternFill("solid", fgColor="E2F0D9")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red    = PatternFill("solid", fgColor="FDE9D9")
    wh_b   = Font(color="FFFFFF", bold=True, size=11)
    bld    = Font(bold=True)
    thin   = Side(style="thin", color="BFBFBF")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    dfmt   = "DD/MM/YYYY"

    # ── Sheet 1: Scenario Result ────────────────────────────────
    hdrs = [
        "SKU", "Description", "Market", "Product Group", "Fixture",
        "Priority", "Weekly Cap", "Ship Request", "Demand Qty",
    ]
    for _, d in date_cols:
        hdrs.append(d.strftime("%d/%m"))
    hdrs += ["Beg Inv", "Period Prod", "Available",
             "End Inv", "Shortage", "Risk", "Horizon Status"]

    for c, h in enumerate(hdrs, 1):
        cell = ws_out.cell(1, c, h)
        cell.fill = teal
        cell.font = wh_b
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center", wrap_text=True)
        cell.border = bdr

    tot_plan = tot_short = 0
    for rx, rd in enumerate(rows, 2):
        i = rd["row"]
        s = rd["sku"]
        periods = sku_groups[s]
        j = next(k for k, p in enumerate(periods) if p["row"] == i)
        pri = priority_map.get(s, 2)

        for c, v in enumerate([
            rd["sku"], rd["description"], rd["market"],
            rd["product_group"], rd["fixture"], pri,
            rd["weekly_cap"], rd["ship_request"], rd["demand"],
        ], 1):
            ws_out.cell(rx, c).value = v
        ws_out.cell(rx, 8).number_format = dfmt

        ship      = rd["ship_request"]
        prev_ship = periods[j - 1]["ship_request"] if j > 0 else None

        pp = 0
        for t, (_, d) in enumerate(date_cols):
            dv = solve_as_int(value(prod[(s, t)]))
            in_per = ((prev_ship is None and d <= ship)
                      or (prev_ship is not None and d > prev_ship and d <= ship))
            ws_out.cell(rx, 10 + t).value = dv if in_per else 0
            if in_per:
                pp += dv

        beg = (INITIAL_FG_INV if j == 0
               else solve_as_int(value(inv[periods[j - 1]["row"]])))
        ei  = solve_as_int(value(inv[i]))
        sh  = solve_as_int(value(short[i]))
        avl = beg + pp

        sc = 10 + num_days
        ws_out.cell(rx, sc).value     = beg
        ws_out.cell(rx, sc + 1).value = pp
        ws_out.cell(rx, sc + 2).value = avl
        ws_out.cell(rx, sc + 3).value = ei
        ws_out.cell(rx, sc + 4).value = sh

        tot_plan  += pp
        tot_short += sh

        risk = "SHORTAGE" if sh > 0 else "OK"
        ws_out.cell(rx, sc + 5).value = risk

        last_d = date_cols[-1][1] if date_cols else None
        hs = ("WITHIN_HORIZON"
              if ship and last_d and ship <= last_d
              else "BEYOND_HORIZON")
        ws_out.cell(rx, sc + 6).value = hs

        rc = ws_out.cell(rx, sc + 5)
        rc.fill = green if risk == "OK" else red
        if risk != "OK":
            rc.font = bld

        for c in range(1, len(hdrs) + 1):
            ws_out.cell(rx, c).border = bdr

    for col_cells in ws_out.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_out.column_dimensions[cl].width = min(max(mx + 2, 8), 22)
    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions

    # ── Sheet 2: KPI Summary ───────────────────────────────────
    ws_kpi = wb_out.create_sheet("KPI_Summary")
    kpi = [
        ["Metric", "Value"],
        ["Solver Status", LpStatus[model.status]],
        ["Total Demand Lines", len(rows)],
        ["Planning Horizon (workdays)", num_days],
        ["Total Demand Qty", sum(rd["demand"] for rd in rows)],
        ["Total Planned Qty", tot_plan],
        ["Total Shortage Qty", tot_short],
        ["Lines with Shortage",
         sum(1 for rd in rows if solve_as_int(value(short[rd["row"]])) > 0)],
        ["Lines OK",
         sum(1 for rd in rows if solve_as_int(value(short[rd["row"]])) == 0)],
    ]
    for ri, rdata in enumerate(kpi, 1):
        for ci, v in enumerate(rdata, 1):
            c = ws_kpi.cell(ri, ci, v)
            c.border = bdr
    for c in range(1, 3):
        c2 = ws_kpi.cell(1, c)
        c2.fill, c2.font = teal, wh_b
        c2.alignment = Alignment(horizontal="center")
    for ri in range(2, len(kpi) + 1):
        ws_kpi.cell(ri, 1).font = bld
    ws_kpi.column_dimensions["A"].width = 30
    ws_kpi.column_dimensions["B"].width = 18

    # ── Sheet 3: SKU Inventory Flow ────────────────────────────
    ws_iv = wb_out.create_sheet("SKU_Inventory_Flow")
    ivh = ["SKU", "Priority", "Period", "Ship Date", "Beg Inv",
           "Period Prod", "Available", "Demand", "End Inv",
           "Shortage", "Short %", "Risk"]
    for c, h in enumerate(ivh, 1):
        cell = ws_iv.cell(1, c, h)
        cell.fill, cell.font = teal, wh_b
        cell.alignment = Alignment(horizontal="center")
        cell.border = bdr
    ir = 2
    for s in all_skus:
        pri = priority_map.get(s, 2)
        for j, rd in enumerate(sku_groups[s]):
            i    = rd["row"]
            ship = rd["ship_request"]
            ps   = sku_groups[s][j - 1]["ship_request"] if j > 0 else None
            pp2  = sum(
                solve_as_int(value(prod[(s, t)]))
                for t, (_, d) in enumerate(date_cols)
                if (ps is None and d <= ship) or (ps and d > ps and d <= ship)
            )
            bg = (INITIAL_FG_INV if j == 0
                  else solve_as_int(value(inv[sku_groups[s][j-1]["row"]])))
            ei = solve_as_int(value(inv[i]))
            sh = solve_as_int(value(short[i]))
            av = bg + pp2
            rk = "SHORTAGE" if sh > 0 else "OK"
            dem = rd["demand"]
            sp = f"{100*sh/dem:.1f}%" if dem > 0 else "0.0%"
            for ci, v in enumerate([s, pri, j+1, ship, bg, pp2, av,
                                     dem, ei, sh, sp, rk], 1):
                cl = ws_iv.cell(ir, ci)
                cl.value  = v
                cl.border = bdr
            ws_iv.cell(ir, 4).number_format = dfmt
            rc2 = ws_iv.cell(ir, 12)
            rc2.fill = green if rk == "OK" else red
            if rk != "OK":
                rc2.font = bld
            ir += 1
    for col_cells in ws_iv.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_iv.column_dimensions[cl].width = min(max(mx + 2, 10), 20)

    # ── Sheet 4: Fixture Load Detail ───────────────────────────
    ws_fix = wb_out.create_sheet("Fixture_Load_Detail")
    fh = ["Fixture", "Daily Cap"] + [d.strftime("%d/%m") for _, d in date_cols]
    for c, h in enumerate(fh, 1):
        cell = ws_fix.cell(1, c, h)
        cell.fill, cell.font = teal, wh_b
        cell.alignment = Alignment(horizontal="center")
        cell.border = bdr
    for fi, (fx, dc) in enumerate(sorted(fixture_daily_cap.items()), 2):
        ws_fix.cell(fi, 1).value = fx
        ws_fix.cell(fi, 1).font  = bld
        ws_fix.cell(fi, 2).value = round(dc, 1)
        for c in (1, 2):
            ws_fix.cell(fi, c).border = bdr
        sfx = [s for s in all_skus if sku_fixture.get(s) == fx]
        for t in range(num_days):
            ld = sum(solve_as_int(value(prod[(s, t)])) for s in sfx)
            cl = ws_fix.cell(fi, 3 + t)
            cl.value  = ld
            cl.border = bdr
            if ld > dc + 0.5:
                cl.fill, cl.font = red, bld
            elif ld > dc * 0.9:
                cl.fill = yellow

    for col_cells in ws_fix.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_fix.column_dimensions[cl].width = min(max(mx + 2, 8), 16)
    ws_fix.freeze_panes = "C2"

    # ── Sheet 5: MPS Comparison ─────────────────────────────────
    ws_cmp = wb_out.create_sheet("MPS_Comparison")
    cmp_hdrs = ["SKU", "Priority", "Fixture",
                "MPS Demand", "MPS Planned", "MPS Shortage", "MPS Short%",
                "PP Demand", "PP Planned", "PP Shortage", "PP Short%",
                "Gap (PP-MPS)"]
    for c, h in enumerate(cmp_hdrs, 1):
        cell = ws_cmp.cell(1, c, h)
        cell.fill, cell.font = teal, wh_b
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bdr

    # Read MPS data for horizon
    mps_df = pd.read_excel(INPUT_FILE, sheet_name="MPS", header=None)
    mps_agg = defaultdict(lambda: {"dem": 0, "plan": 0, "short": 0})
    for mr in range(1, mps_df.shape[0]):
        msku = mps_df.iloc[mr, 0]
        mwk  = pd.to_datetime(mps_df.iloc[mr, 2], errors="coerce")
        if pd.isna(msku) or pd.isna(mwk):
            continue
        if mwk < pd.Timestamp("2026-04-13") or mwk > pd.Timestamp("2026-05-25"):
            continue
        msku = str(msku).strip()
        mps_agg[msku]["dem"]   += int(float(mps_df.iloc[mr, 4])) if pd.notna(mps_df.iloc[mr, 4]) else 0
        mps_agg[msku]["plan"]  += int(float(mps_df.iloc[mr, 9])) if pd.notna(mps_df.iloc[mr, 9]) else 0
        mps_agg[msku]["short"] += int(float(mps_df.iloc[mr, 11])) if pd.notna(mps_df.iloc[mr, 11]) else 0

    # Solver aggregation
    pp_agg = defaultdict(lambda: {"dem": 0, "plan": 0, "short": 0})
    for rd in rows:
        s = rd["sku"]
        i = rd["row"]
        pp_agg[s]["dem"]   += rd["demand"]
        pp_agg[s]["plan"]  += sum(
            solve_as_int(value(prod[(s, t)]))
            for t in range(num_days)
        ) if s not in pp_agg or pp_agg[s]["plan"] == 0 else 0
        pp_agg[s]["short"] += solve_as_int(value(short[i]))

    # Recalculate planned properly by SKU
    for s in all_skus:
        pp_agg[s]["plan"] = sum(
            solve_as_int(value(prod[(s, t)])) for t in range(num_days)
        )

    cmp_row = 2
    all_cmp = sorted(set(list(mps_agg.keys()) + list(pp_agg.keys())))
    t_md = t_mp = t_ms = t_pd = t_pp = t_ps = 0
    for s in all_cmp:
        md = mps_agg[s]["dem"]
        mp = mps_agg[s]["plan"]
        ms = mps_agg[s]["short"]
        mr_pct = f"{100*ms/md:.1f}%" if md else "0.0%"
        pd2 = pp_agg[s]["dem"]
        pp2 = pp_agg[s]["plan"]
        ps = pp_agg[s]["short"]
        pr_pct = f"{100*ps/pd2:.1f}%" if pd2 else "0.0%"
        gap = ps - ms
        pri = priority_map.get(s, "?")
        fx = sku_fixture.get(s, "?")
        for ci, v in enumerate([s, pri, fx, md, mp, ms, mr_pct,
                                 pd2, pp2, ps, pr_pct, gap], 1):
            cl = ws_cmp.cell(cmp_row, ci)
            cl.value = v
            cl.border = bdr
        # Color gap
        gap_cell = ws_cmp.cell(cmp_row, 12)
        if gap > 0:
            gap_cell.fill = red
            gap_cell.font = bld
        elif gap < 0:
            gap_cell.fill = green
        t_md += md; t_mp += mp; t_ms += ms
        t_pd += pd2; t_pp += pp2; t_ps += ps
        cmp_row += 1

    # Totals row
    for ci, v in enumerate(["TOTAL", "", "", t_md, t_mp, t_ms,
                             f"{100*t_ms/t_md:.1f}%" if t_md else "",
                             t_pd, t_pp, t_ps,
                             f"{100*t_ps/t_pd:.1f}%" if t_pd else "",
                             t_ps - t_ms], 1):
        cl = ws_cmp.cell(cmp_row, ci)
        cl.value = v
        cl.font = bld
        cl.border = bdr

    for col_cells in ws_cmp.columns:
        cl2 = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_cmp.column_dimensions[cl2].width = min(max(mx + 3, 12), 20)

    # ── Save & report ───────────────────────────────────────────
    wb_out.save(OUTPUT_FILE)

    # Summary by priority
    pri1_short = sum(solve_as_int(value(short[rd["row"]])) for rd in rows
                     if priority_map.get(rd["sku"], 2) == 1)
    pri1_dem   = sum(rd["demand"] for rd in rows
                     if priority_map.get(rd["sku"], 2) == 1)
    pri2_short = sum(solve_as_int(value(short[rd["row"]])) for rd in rows
                     if priority_map.get(rd["sku"], 2) == 2)
    pri2_dem   = sum(rd["demand"] for rd in rows
                     if priority_map.get(rd["sku"], 2) == 2)

    print()
    print("=" * 60)
    print("  BASELINE PLAN (before scenario)")
    print("=" * 60)
    print(f"  Solver status    : {LpStatus[model.status]}")
    print(f"  Demand lines     : {len(rows)}")
    print(f"  Planning horizon : {num_days} workdays")
    print(f"  Total demand     : {sum(rd['demand'] for rd in rows):,}")
    print(f"  Total planned    : {tot_plan:,}")
    print(f"  Total shortage   : {tot_short:,}")
    print("-" * 60)
    if pri1_dem:
        print(f"  Pri-1 (A,C,E) shortage: {pri1_short:>8,} / {pri1_dem:>8,}  "
              f"({100*pri1_short/pri1_dem:.1f}%)")
    if pri2_dem:
        print(f"  Pri-2 (B,D,F) shortage: {pri2_short:>8,} / {pri2_dem:>8,}  "
              f"({100*pri2_short/pri2_dem:.1f}%)")
    print("-" * 60)
    print("  MPS vs Baseline comparison:")
    print(f"    MPS  shortage: {t_ms:>8,} ({100*t_ms/t_md:.1f}%)")
    print(f"    PP   shortage: {t_ps:>8,} ({100*t_ps/t_pd:.1f}%)")
    print(f"    Gap         : {t_ps - t_ms:>+8,}")
    print("-" * 60)
    print(f"  Output: {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()