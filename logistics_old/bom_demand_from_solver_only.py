
from pathlib import Path
import math
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Anchor all paths to the folder containing THIS script ──────────────
BASE_DIR = Path(__file__).resolve().parent

INPUT_FILE = BASE_DIR / "solver.xlsx"
OUTPUT_EXCEL = BASE_DIR / "transport_output_from_solver_only.xlsx"
CSV_US = BASE_DIR / "transport_us_from_solver.csv"
CSV_UK = BASE_DIR / "transport_uk_from_solver.csv"
CSV_AU = BASE_DIR / "transport_au_from_solver.csv"
CSV_ALL = BASE_DIR / "transport_weekly_all_markets_from_solver.csv"
CSV_SUMMARY_MARKET = BASE_DIR / "transport_summary_market_from_solver.csv"
CSV_SUMMARY_ALL = BASE_DIR / "transport_summary_all_from_solver.csv"

# ── Partial container loading ──────────────────────────────────────────
# Containers do NOT need to be 100% full.
# fill_rate = 0.80 means each container can hold at most 80% of nominal CBM.
# Cost is still the FULL container price (you pay for the whole box).
FILL_RATE_40 = 0.80   # 40-ft effective fill rate
FILL_RATE_20 = 0.80   # 20-ft effective fill rate

def load_input(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb["solver"]

def parse_solver_sheet(ws):
    sku_master = []
    r = 6
    while True:
        item = ws.cell(r, 1).value
        if item in (None, ""):
            break
        sku_master.append({
            "Item": str(item).strip(),
            "Des": ws.cell(r, 2).value,
            "PackSize": float(ws.cell(r, 3).value),
            "CBM100": float(ws.cell(r, 4).value),
            "Price": float(ws.cell(r, 5).value),
            "Market": str(ws.cell(r, 6).value).strip(),
        })
        r += 1

    lanes = []
    r = 15
    while True:
        lane_key = ws.cell(r, 5).value
        if lane_key in (None, ""):
            break
        lane_key = str(lane_key).strip()
        market = ws.cell(r, 1).value
        if market not in ["US", "UK", "AU"]:
            if lane_key.startswith("US"):
                market = "US"
            elif lane_key.startswith("UK"):
                market = "UK"
            elif lane_key.startswith("AU"):
                market = "AU"
        lanes.append({
            "Market": market,
            "Mode": str(ws.cell(r, 2).value).strip(),
            "Cost": float(ws.cell(r, 3).value),
            "Cap_CBM": float(ws.cell(r, 4).value),
            "LaneKey": lane_key,
        })
        r += 1

    item_headers = []
    c = 2
    while True:
        h = ws.cell(27, c).value
        if h in (None, ""):
            break
        item_headers.append(str(h).strip())
        c += 1

    decisions = []
    r = 28
    while True:
        week = ws.cell(r, 1).value
        if week in (None, ""):
            break
        rec = {"Week": week}
        for i, item in enumerate(item_headers, start=2):
            rec[f"Units_{item}"] = float(ws.cell(r, i).value or 0)
        decisions.append(rec)
        r += 1

    return pd.DataFrame(sku_master), pd.DataFrame(lanes), pd.DataFrame(decisions)

def solve_week(demand_cbm, lp):
    """Find cheapest mix of 40-ft, 20-ft containers (partially loadable) + LCL."""
    if demand_cbm <= 1e-12:
        return {"n40": 0, "n20": 0, "LCL_CBM": 0.0,
                "Coverage_CBM": 0.0, "Slack_CBM": 0.0,
                "Cost": 0.0, "Util_Pct": 0.0}

    # Containers can hold up to their nominal capacity. They don't have to be 100% full.
    eff_cap40 = lp["cap40"]
    eff_cap20 = lp["cap20"]

    max40 = math.ceil(demand_cbm / eff_cap40) + 2 if eff_cap40 > 0 else 0
    max20 = math.ceil(demand_cbm / eff_cap20) + 2 if eff_cap20 > 0 else 0
    best = None
    for n40 in range(max40 + 1):
        for n20 in range(max20 + 1):
            # Total capacity provided by rented containers
            fixed_cov = n40 * eff_cap40 + n20 * eff_cap20
            # Whatever exceeds container capacity goes to LCL
            lcl = max(0.0, demand_cbm - fixed_cov)
            
            coverage = fixed_cov + lcl
            slack = coverage - demand_cbm  # The unused space inside the containers
            
            # Cost: full container price + LCL per-CBM
            cost = n40 * lp["cost40"] + n20 * lp["cost20"] + lcl * lp["cost_lcl"]
            cand = (round(cost, 8), n40 + n20, round(lcl, 8),
                    round(slack, 8), n40, n20, lcl, coverage, slack, cost)
            if best is None or cand < best:
                best = cand
    _, _, _, _, n40, n20, lcl, coverage, slack, cost = best

    # Container utilization %
    total_eff_cap = n40 * eff_cap40 + n20 * eff_cap20
    used_in_cont = min(demand_cbm, total_eff_cap)
    util_pct = (used_in_cont / total_eff_cap * 100.0) if total_eff_cap > 0 else 0.0

    return {
        "n40": int(n40),
        "n20": int(n20),
        "LCL_CBM": float(lcl),
        "Coverage_CBM": float(coverage),
        "Slack_CBM": float(slack),
        "Cost": float(cost),
        "Util_Pct": round(util_pct, 1),
    }

def build_outputs(sku_df, lanes_df, decisions_df):
    sku_lookup = sku_df.set_index("Item").to_dict("index")
    markets = ["US", "UK", "AU"]
    market_items = {m: sku_df.loc[sku_df["Market"] == m, "Item"].tolist() for m in markets}
    lane_params = {}
    for m in markets:
        sub = lanes_df[lanes_df["Market"] == m].set_index("Mode")
        lane_params[m] = {
            "cost40": float(sub.loc["40", "Cost"]),
            "cap40": float(sub.loc["40", "Cap_CBM"]),
            "cost20": float(sub.loc["20", "Cost"]),
            "cap20": float(sub.loc["20", "Cap_CBM"]),
            "cost_lcl": float(sub.loc["LCL", "Cost"]),
            "fill40": FILL_RATE_40,
            "fill20": FILL_RATE_20,
        }

    market_dfs = {}
    all_rows = []
    summary_market_rows = []

    for m in markets:
        rows = []
        items = market_items[m]
        for _, row in decisions_df.iterrows():
            rec = {
                "Market": m,
                "Week": row["Week"],
                "Week_Date": row["Week"].strftime("%Y-%m-%d") if hasattr(row["Week"], "strftime") else str(row["Week"])
            }
            total_cbm = 0.0
            for item in items:
                units = float(row.get(f"Units_{item}", 0) or 0)
                pack = float(sku_lookup[item]["PackSize"])
                cbm100 = float(sku_lookup[item]["CBM100"])
                boxes = math.ceil(units / pack) if units > 0 else 0
                cbm = boxes * cbm100 / 100.0
                rec[f"Units_{item}"] = int(units) if units.is_integer() else units
                rec[f"Boxes_{item}"] = int(boxes)
                rec[f"CBM_{item}"] = cbm
                total_cbm += cbm

            rec["Total_Demand_CBM"] = total_cbm
            rec.update(solve_week(total_cbm, lane_params[m]))
            rows.append(rec)
            all_rows.append(rec.copy())

        df = pd.DataFrame(rows)
        ordered = ["Week", "Week_Date"]
        for prefix in ["Units", "Boxes", "CBM"]:
            ordered += [f"{prefix}_{item}" for item in items]
        ordered += ["Total_Demand_CBM", "n40", "n20", "LCL_CBM",
                    "Coverage_CBM", "Slack_CBM", "Util_Pct", "Cost"]
        df = df[ordered]
        market_dfs[m] = df

        summary_market_rows.append({
            "Market": m,
            "Weeks": len(df),
            "Total_Demand_CBM": df["Total_Demand_CBM"].sum(),
            "Total_n40": df["n40"].sum(),
            "Total_n20": df["n20"].sum(),
            "Total_LCL_CBM": df["LCL_CBM"].sum(),
            "Total_Coverage_CBM": df["Coverage_CBM"].sum(),
            "Total_Slack_CBM": df["Slack_CBM"].sum(),
            "Avg_Util_Pct": df.loc[df["Util_Pct"] > 0, "Util_Pct"].mean() if (df["Util_Pct"] > 0).any() else 0.0,
            "Total_Cost": df["Cost"].sum(),
            "Avg_Cost_per_CBM": (df["Cost"].sum() / df["Total_Demand_CBM"].sum()) if df["Total_Demand_CBM"].sum() else 0
        })

    all_df = pd.DataFrame(all_rows).sort_values(["Week", "Market"]).reset_index(drop=True)
    summary_market_df = pd.DataFrame(summary_market_rows)
    summary_all_df = pd.DataFrame([{
        "Weeks": len(decisions_df),
        "Markets": len(markets),
        "Grand_Total_Demand_CBM": summary_market_df["Total_Demand_CBM"].sum(),
        "Grand_Total_n40": summary_market_df["Total_n40"].sum(),
        "Grand_Total_n20": summary_market_df["Total_n20"].sum(),
        "Grand_Total_LCL_CBM": summary_market_df["Total_LCL_CBM"].sum(),
        "Grand_Total_Coverage_CBM": summary_market_df["Total_Coverage_CBM"].sum(),
        "Grand_Total_Slack_CBM": summary_market_df["Total_Slack_CBM"].sum(),
        "Grand_Total_Cost": summary_market_df["Total_Cost"].sum(),
        "Avg_Cost_per_CBM": (summary_market_df["Total_Cost"].sum()/summary_market_df["Total_Demand_CBM"].sum())
            if summary_market_df["Total_Demand_CBM"].sum() else 0
    }])
    return market_dfs, all_df, summary_market_df, summary_all_df

def write_excel(market_dfs, all_df, summary_market_df, summary_all_df, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet_from_df(name, df):
        ws = wb.create_sheet(title=name)
        ws.append(df.columns.tolist())
        for row in df.itertuples(index=False):
            ws.append(list(row))
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:100])
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 24)
        header_map = {cell.column: cell.value for cell in ws[1]}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                h = header_map[cell.column]
                if h == "Week":
                    cell.number_format = "yyyy-mm-dd"
                elif h in ["Total_Demand_CBM", "LCL_CBM", "Coverage_CBM", "Slack_CBM"] or str(h).startswith("CBM_"):
                    cell.number_format = "0.0000"
                elif "Cost" in str(h) or "Avg_Cost" in str(h):
                    cell.number_format = "$#,##0.00"
                elif str(h).startswith("Units_") or str(h).startswith("Boxes_") or h in ["n40", "n20", "Weeks", "Markets"]:
                    cell.number_format = "0"

    add_sheet_from_df("US", market_dfs["US"])
    add_sheet_from_df("UK", market_dfs["UK"])
    add_sheet_from_df("AU", market_dfs["AU"])
    add_sheet_from_df("Weekly_All_Markets", all_df)
    add_sheet_from_df("Summary_Market", summary_market_df)
    add_sheet_from_df("Summary_All", summary_all_df)
    wb.save(out_path)

def main():
    print("=" * 60)
    print("  BOM Demand -> Transport Solver  (partial loading OK)")
    print("=" * 60)
    print(f"  Fill Rate:  40-ft = {FILL_RATE_40*100:.0f}%  |  20-ft = {FILL_RATE_20*100:.0f}%")
    print(f"  Input :  {INPUT_FILE}")
    print(f"  Output:  {OUTPUT_EXCEL}")
    print()

    ws = load_input(INPUT_FILE)
    sku_df, lanes_df, decisions_df = parse_solver_sheet(ws)
    market_dfs, all_df, summary_market_df, summary_all_df = build_outputs(sku_df, lanes_df, decisions_df)

    market_dfs["US"].to_csv(CSV_US, index=False)
    market_dfs["UK"].to_csv(CSV_UK, index=False)
    market_dfs["AU"].to_csv(CSV_AU, index=False)
    all_df.to_csv(CSV_ALL, index=False)
    summary_market_df.to_csv(CSV_SUMMARY_MARKET, index=False)
    summary_all_df.to_csv(CSV_SUMMARY_ALL, index=False)
    write_excel(market_dfs, all_df, summary_market_df, summary_all_df, OUTPUT_EXCEL)

    # -- Console report ---------------------------------------------
    print("\n-- Summary by Market " + "-" * 39)
    print(summary_market_df.to_string(index=False))
    print("\n-- Overall Summary " + "-" * 41)
    print(summary_all_df.to_string(index=False))
    print("\n[OK] Done. Files written.")

if __name__ == "__main__":
    main()
