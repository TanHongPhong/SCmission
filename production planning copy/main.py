"""
Scenario Production Plan -- PCBA Firmware Defect (SKU A)
========================================================
Recovery level: 60% (A shortage capped at 11,674 FG)

Disruption:
  - SKU A PCBA (RM04) has firmware defect -> all stock ON HOLD
  - Reflash: equipment arrives in 3 weeks (May 6)
  - New supply: 4000 PCBA/day, air (3d, +$2) or sea (21d)

Model constraints:
  1. Fixture capacity (shared, daily)
  2. Inventory continuity per SKU
  3. Priority allocation (A>B, C>D, E>F)
  4. PCBA cumulative supply >= cumulative A production x BOM
  5. SKU A total shortage <= MAX_A_SHORTAGE (60% recovery)

Compare with audit_current_plan.py (baseline, no disruption).
"""

import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pulp import (
    LpProblem, LpMinimize, LpVariable, lpSum, LpInteger,
    PULP_CBC_CMD, LpStatus, value
)
from datetime import datetime, date, timedelta
from collections import defaultdict
import pandas as pd

INPUT_FILE  = "SCM_round2.1_new.xlsx"
OUTPUT_FILE = "SCM_round2.1_new_solved.xlsx"
PLAN_SHEET  = "Production_plan"
INPUT_SHEET = "input"

# ── Penalty weights ─────────────────────────────────────────
M1_PRI1 = 1_000_000     # Priority-1 shortage penalty
M1_PRI2 = 10_000         # Priority-2 shortage penalty
M2      = 0.001          # Inventory holding (tiny, prefer JIT)

# ── Scenario: PCBA firmware defect ──────────────────────────
PCBA_SKU         = "A"
PCBA_BOM         = 2             # 2 PCBA per FG
PCBA_ON_HOLD     = 370_000       # 340K stock + 30K PO, all defective
REFLASH_WAIT     = 21            # 3 weeks for equipment (calendar days)
SUPPLIER_CAP     = 4_000         # new PCBA/day from supplier
AIR_LEAD         = 3             # air freight (calendar days)
SEA_LEAD         = 21            # sea freight (calendar days)
AIR_COST         = 2.0           # USD extra per PCBA shipped by air

# ── Recovery target ─────────────────────────────────────────
# From recovery_analysis.py:
#   Worst A shortage (0%) = 29,184      Best (100%) = 0
#   60% recovery -> A shortage <= 11,674
RECOVERY_PCT     = 60
MAX_A_SHORTAGE   = 11_674        # = 29184 - 0.60 * 29184

INITIAL_FG_INV = 0


# ── Helpers ─────────────────────────────────────────────────

def to_date(x):
    if isinstance(x, datetime): return x.date()
    if isinstance(x, date):     return x
    return None

def num(x, default=0):
    if x is None or str(x).strip() == "": return default
    try:    return float(x)
    except: return default

def solve_as_int(v):
    return int(round(v)) if v is not None else 0

def read_priorities(f, s):
    df = pd.read_excel(f, sheet_name=s, header=None)
    m = {}
    for r in range(df.shape[0]):
        sku, pri = df.iloc[r, 0], df.iloc[r, 6]
        if pd.notna(sku) and pd.notna(pri) and str(sku).strip() not in ('SKU', ''):
            try: m[str(sku).strip()] = int(pri)
            except: pass
    return m


# ── Main ────────────────────────────────────────────────────

def main():
    # ── 1. Read input ───────────────────────────────────────
    priority_map = read_priorities(INPUT_FILE, INPUT_SHEET)
    print("[INFO] Priorities:", {k: f"P{v}" for k, v in sorted(priority_map.items())})

    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb[PLAN_SHEET]

    date_cols = []
    col = 11
    while True:
        try:    d = to_date(ws.cell(42, col).value)
        except: d = None
        if d is None: break
        date_cols.append((col, d))
        col += 1
    num_days = len(date_cols)
    horizon_start = date_cols[0][1]
    horizon_end   = date_cols[-1][1]
    print(f"[INFO] {num_days} workdays ({horizon_start} to {horizon_end})")

    rows = []
    r = 43
    while True:
        sku = ws.cell(r, 1).value
        if sku is None or str(sku).strip() == "": break
        rows.append({
            "row": r,
            "sku":          str(sku).strip(),
            "description":  str(ws.cell(r, 2).value or ""),
            "market":       str(ws.cell(r, 3).value or ""),
            "product_group":str(ws.cell(r, 4).value or ""),
            "fixture":      str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else "N/A",
            "weekly_cap":   float(num(ws.cell(r, 6).value, 0)),
            "ship_request": to_date(ws.cell(r, 7).value),
            "demand":       int(num(ws.cell(r, 8).value, 0)),
        })
        r += 1
    print(f"[INFO] {len(rows)} demand lines")

    # ── 2. Group & capacities ───────────────────────────────
    sku_groups  = defaultdict(list)
    sku_fixture = {}
    for rd in rows:
        sku_groups[rd["sku"]].append(rd)
        sku_fixture[rd["sku"]] = rd["fixture"]
    for s in sku_groups:
        sku_groups[s].sort(key=lambda x: x["ship_request"])
    all_skus = sorted(sku_groups.keys())

    fixture_max_cap = {}
    for rd in rows:
        fx = rd["fixture"]
        if fx != "N/A":
            if fx not in fixture_max_cap or rd["weekly_cap"] > fixture_max_cap[fx]:
                fixture_max_cap[fx] = rd["weekly_cap"]
    fixture_daily_cap = {fx: cap / 6.0 for fx, cap in fixture_max_cap.items()}

    print("[INFO] Fixture caps:", {fx: f"{c:.0f}/day" for fx, c in fixture_daily_cap.items()})

    # ── 3. Build LP ─────────────────────────────────────────
    model = LpProblem("Scenario_PCBA", LpMinimize)

    # Production per SKU per day
    prod = {}
    for s in all_skus:
        for t in range(num_days):
            prod[(s, t)] = LpVariable(f"p_{s}_d{t}", 0, cat=LpInteger)

    # Inventory & shortage per demand-period
    inv   = {}
    short = {}
    for rd in rows:
        i = rd["row"]
        inv[i]   = LpVariable(f"inv_r{i}",   0, cat=LpInteger)
        short[i] = LpVariable(f"short_r{i}", 0, cat=LpInteger)

    # ── PCBA logistics: supplier air/sea per calendar day ───
    max_cal = (horizon_end - horizon_start).days          # ~30
    air_ship = {}
    sea_ship = {}
    for d in range(max_cal + 1):
        air_ship[d] = LpVariable(f"air_{d}", 0, cat=LpInteger)
        sea_ship[d] = LpVariable(f"sea_{d}", 0, cat=LpInteger)
        model += air_ship[d] + sea_ship[d] <= SUPPLIER_CAP

    # ── Objective ───────────────────────────────────────────
    # PRIMARY: minimize air freight cost
    # SECONDARY: small shortage penalty for priority allocation tie-breaking
    # The A shortage cap (Constraint D) enforces the service level
    obj = [AIR_COST * lpSum(air_ship[d] for d in range(max_cal + 1))]
    for rd in rows:
        pri = priority_map.get(rd["sku"], 2)
        pen = 10 if pri == 1 else 1    # small, just for allocation priority
        obj.append(0.1 * pen * short[rd["row"]])
    obj.append(M2 * lpSum(inv[rd["row"]] for rd in rows))
    model += lpSum(obj)

    # ── Constraint A: fixture daily capacity ────────────────
    for fx, dc in fixture_daily_cap.items():
        if fx == "N/A" or dc <= 0: continue
        sfx = [s for s in all_skus if sku_fixture.get(s) == fx]
        for t in range(num_days):
            model += lpSum(prod[(s, t)] for s in sfx) <= dc

    # ── Constraint B: inventory balance per SKU ─────────────
    for s, periods in sku_groups.items():
        for j, rd in enumerate(periods):
            i    = rd["row"]
            ship = rd["ship_request"]
            if j == 0:
                beg = INITIAL_FG_INV
                pp  = lpSum(prod[(s, t)] for t, (_, d) in enumerate(date_cols) if d <= ship)
            else:
                beg = inv[periods[j-1]["row"]]
                ps  = periods[j-1]["ship_request"]
                pp  = lpSum(prod[(s, t)] for t, (_, d) in enumerate(date_cols)
                            if d > ps and d <= ship)
            model += inv[i] - short[i] == beg + pp - rd["demand"]

    # ── Constraint C: PCBA supply for SKU A ─────────────────
    #  cumul_A_prod(t) * BOM <= reflash(t) + cumul_air(t) + cumul_sea(t)
    print(f"[INFO] PCBA scenario: on-hold={PCBA_ON_HOLD:,}, reflash@day {REFLASH_WAIT}, "
          f"supplier={SUPPLIER_CAP}/day, air={AIR_LEAD}d(+${AIR_COST}), sea={SEA_LEAD}d")

    for t, (_, dt) in enumerate(date_cols):
        cal = (dt - horizon_start).days

        # Reflashed stock (all available from day 21)
        reflash = PCBA_ON_HOLD if cal >= REFLASH_WAIT else 0

        # Cumulative new PCBA arrived from air and sea
        cumul_new = (
            lpSum(air_ship[d] for d in range(max_cal + 1) if d + AIR_LEAD <= cal)
          + lpSum(sea_ship[d] for d in range(max_cal + 1) if d + SEA_LEAD <= cal)
        )

        # Cumulative A production up to day t
        cumul_a = lpSum(prod[(PCBA_SKU, t2)] for t2 in range(t + 1))

        model += cumul_a * PCBA_BOM <= reflash + cumul_new

    # ── Constraint D: SKU A shortage cap (recovery level) ───
    a_rows = [rd for rd in rows if rd["sku"] == PCBA_SKU]
    model += lpSum(short[rd["row"]] for rd in a_rows) <= MAX_A_SHORTAGE
    print(f"[INFO] Recovery {RECOVERY_PCT}%: A shortage capped at {MAX_A_SHORTAGE:,}")

    # ── 4. Solve ────────────────────────────────────────────
    print("[INFO] Solving ...")
    solver = PULP_CBC_CMD(msg=False, timeLimit=180)
    model.solve(solver)
    print(f"[INFO] Status: {LpStatus[model.status]}")

    # ── 5. Compute scenario metrics ─────────────────────────
    total_air = sum(solve_as_int(value(air_ship[d])) for d in range(max_cal + 1))
    total_sea = sum(solve_as_int(value(sea_ship[d])) for d in range(max_cal + 1))
    total_air_cost = total_air * AIR_COST

    # ── 6. Build output workbook ────────────────────────────
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Scenario_Result"

    # Styles
    navy   = PatternFill("solid", fgColor="1F4E78")
    orange = PatternFill("solid", fgColor="E65100")
    green  = PatternFill("solid", fgColor="E2F0D9")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red    = PatternFill("solid", fgColor="FDE9D9")
    wh_b   = Font(color="FFFFFF", bold=True, size=11)
    bld    = Font(bold=True)
    thin   = Side(style="thin", color="BFBFBF")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    dfmt   = "DD/MM/YYYY"

    # ── Sheet 1: Production Plan Result ─────────────────────
    hdrs = ["SKU", "Description", "Market", "Product Group", "Fixture",
            "Priority", "Weekly Cap", "Ship Request", "Demand Qty"]
    for _, d in date_cols:
        hdrs.append(d.strftime("%d/%m"))
    hdrs += ["Beg Inv", "Period Prod", "Available",
             "End Inv", "Shortage", "Risk", "Horizon Status"]

    for c, h in enumerate(hdrs, 1):
        cell = ws_out.cell(1, c, h)
        cell.fill, cell.font = orange, wh_b
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr

    tot_plan = tot_short = 0
    for rx, rd in enumerate(rows, 2):
        i = rd["row"]; s = rd["sku"]
        periods = sku_groups[s]
        j = next(k for k, p in enumerate(periods) if p["row"] == i)
        pri = priority_map.get(s, 2)

        for c, v in enumerate([rd["sku"], rd["description"], rd["market"],
                                rd["product_group"], rd["fixture"], pri,
                                rd["weekly_cap"], rd["ship_request"], rd["demand"]], 1):
            ws_out.cell(rx, c).value = v
        ws_out.cell(rx, 8).number_format = dfmt

        ship = rd["ship_request"]
        prev_ship = periods[j - 1]["ship_request"] if j > 0 else None

        pp = 0
        for t, (_, d) in enumerate(date_cols):
            dv = solve_as_int(value(prod[(s, t)]))
            in_p = (prev_ship is None and d <= ship) or (prev_ship and d > prev_ship and d <= ship)
            ws_out.cell(rx, 10 + t).value = dv if in_p else 0
            if in_p: pp += dv

        beg = INITIAL_FG_INV if j == 0 else solve_as_int(value(inv[periods[j-1]["row"]]))
        ei  = solve_as_int(value(inv[i]))
        sh  = solve_as_int(value(short[i]))
        avl = beg + pp

        sc = 10 + num_days
        ws_out.cell(rx, sc).value     = beg
        ws_out.cell(rx, sc+1).value   = pp
        ws_out.cell(rx, sc+2).value   = avl
        ws_out.cell(rx, sc+3).value   = ei
        ws_out.cell(rx, sc+4).value   = sh
        tot_plan += pp; tot_short += sh

        risk = "SHORTAGE" if sh > 0 else "OK"
        ws_out.cell(rx, sc+5).value = risk
        hs = "WITHIN_HORIZON" if ship and ship <= horizon_end else "BEYOND_HORIZON"
        ws_out.cell(rx, sc+6).value = hs

        rc = ws_out.cell(rx, sc+5)
        rc.fill = green if risk == "OK" else red
        if risk != "OK": rc.font = bld
        for c in range(1, len(hdrs)+1): ws_out.cell(rx, c).border = bdr

    for col_cells in ws_out.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_out.column_dimensions[cl].width = min(max(mx + 2, 8), 22)
    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions

    # ── Sheet 2: PCBA Supply Timeline ───────────────────────
    ws_pcba = wb_out.create_sheet("PCBA_Supply_Timeline")
    ph = ["Workday", "Date", "Cal Day",
          "Air Arrived", "Sea Arrived", "Reflash Avail",
          "Cumul PCBA", "A Prod Today", "Cumul A Prod",
          "PCBA Used (cum)", "PCBA Remaining"]
    for c, h in enumerate(ph, 1):
        cell = ws_pcba.cell(1, c, h)
        cell.fill, cell.font = orange, wh_b
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bdr

    cumul_a_total = 0
    for t, (_, dt) in enumerate(date_cols):
        cal = (dt - horizon_start).days
        air_today = sum(solve_as_int(value(air_ship[d]))
                        for d in range(max_cal+1) if d + AIR_LEAD == cal)
        sea_today = sum(solve_as_int(value(sea_ship[d]))
                        for d in range(max_cal+1) if d + SEA_LEAD == cal)
        reflash = PCBA_ON_HOLD if cal >= REFLASH_WAIT else 0
        cumul_air_v = sum(solve_as_int(value(air_ship[d]))
                          for d in range(max_cal+1) if d + AIR_LEAD <= cal)
        cumul_sea_v = sum(solve_as_int(value(sea_ship[d]))
                          for d in range(max_cal+1) if d + SEA_LEAD <= cal)
        cumul_pcba = reflash + cumul_air_v + cumul_sea_v

        a_today = solve_as_int(value(prod[(PCBA_SKU, t)]))
        cumul_a_total += a_today
        pcba_used = cumul_a_total * PCBA_BOM
        pcba_rem  = cumul_pcba - pcba_used

        rw = t + 2
        for ci, v in enumerate([t+1, dt, cal, air_today, sea_today, reflash,
                                 cumul_pcba, a_today, cumul_a_total,
                                 pcba_used, pcba_rem], 1):
            cl = ws_pcba.cell(rw, ci)
            cl.value = v; cl.border = bdr
        ws_pcba.cell(rw, 2).number_format = dfmt
        # Highlight blocked days
        if cumul_pcba == 0:
            for ci in range(1, 12):
                ws_pcba.cell(rw, ci).fill = red

    for col_cells in ws_pcba.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_pcba.column_dimensions[cl].width = min(max(mx + 2, 10), 18)

    # ── Sheet 3: Cost Analysis ──────────────────────────────
    ws_cost = wb_out.create_sheet("Cost_Analysis")
    cost_data = [
        ["PCBA Scenario Cost Analysis", ""],
        ["", ""],
        ["Item", "Value"],
        ["PCBA on hold (defective)", f"{PCBA_ON_HOLD:,}"],
        ["Reflash available from", f"Day {REFLASH_WAIT} ({horizon_start + timedelta(days=REFLASH_WAIT)})"],
        ["Supplier daily capacity", f"{SUPPLIER_CAP:,} PCBA/day"],
        ["", ""],
        ["Logistics Decision", ""],
        ["Total PCBA by AIR", f"{total_air:,}"],
        ["Total PCBA by SEA", f"{total_sea:,}"],
        ["Air lead time", f"{AIR_LEAD} days"],
        ["Sea lead time", f"{SEA_LEAD} days"],
        ["", ""],
        ["Cost Breakdown", ""],
        ["Air freight cost", f"${total_air_cost:,.2f}"],
        ["  = {0:,} PCBA x ${1}/unit".format(total_air, AIR_COST), ""],
        ["Sea freight extra cost", "$0.00"],
        ["Reflash cost", "$0.00 (equipment only)"],
        ["", ""],
        ["TOTAL SCENARIO COST", f"${total_air_cost:,.2f}"],
        ["", ""],
        ["Production Impact", ""],
        ["Total demand", f"{sum(rd['demand'] for rd in rows):,}"],
        ["Total planned", f"{tot_plan:,}"],
        ["Total shortage", f"{tot_short:,}"],
        ["SKU A demand", f"{sum(rd['demand'] for rd in rows if rd['sku']=='A'):,}"],
        ["SKU A planned", f"{sum(solve_as_int(value(prod[('A',t)])) for t in range(num_days)):,}"],
    ]
    for ri, (k, v) in enumerate(cost_data, 1):
        ws_cost.cell(ri, 1).value = k
        ws_cost.cell(ri, 2).value = v
        ws_cost.cell(ri, 1).border = bdr
        ws_cost.cell(ri, 2).border = bdr
        if k and not v:
            ws_cost.cell(ri, 1).font = bld
    ws_cost.cell(1, 1).fill = orange
    ws_cost.cell(1, 1).font = wh_b
    # Highlight total cost
    for ri, (k, _) in enumerate(cost_data, 1):
        if k == "TOTAL SCENARIO COST":
            ws_cost.cell(ri, 1).font = Font(bold=True, size=12)
            ws_cost.cell(ri, 2).font = Font(bold=True, size=12, color="E65100")
    ws_cost.column_dimensions["A"].width = 35
    ws_cost.column_dimensions["B"].width = 25

    # ── Sheet 4: KPI Summary ───────────────────────────────
    ws_kpi = wb_out.create_sheet("KPI_Summary")
    pri1_short = sum(solve_as_int(value(short[rd["row"]])) for rd in rows
                     if priority_map.get(rd["sku"], 2) == 1)
    pri1_dem   = sum(rd["demand"] for rd in rows if priority_map.get(rd["sku"], 2) == 1)
    pri2_short = sum(solve_as_int(value(short[rd["row"]])) for rd in rows
                     if priority_map.get(rd["sku"], 2) == 2)
    pri2_dem   = sum(rd["demand"] for rd in rows if priority_map.get(rd["sku"], 2) == 2)
    kpi = [
        ["Metric", "Value"],
        ["Solver Status", LpStatus[model.status]],
        ["Total Demand Lines", len(rows)],
        ["Planning Horizon", f"{num_days} workdays"],
        ["Total Demand Qty", sum(rd["demand"] for rd in rows)],
        ["Total Planned Qty", tot_plan],
        ["Total Shortage Qty", tot_short],
        ["Pri-1 Shortage", f"{pri1_short:,} / {pri1_dem:,} ({100*pri1_short/pri1_dem:.1f}%)" if pri1_dem else "0"],
        ["Pri-2 Shortage", f"{pri2_short:,} / {pri2_dem:,} ({100*pri2_short/pri2_dem:.1f}%)" if pri2_dem else "0"],
        ["Air Freight PCBA", total_air],
        ["Sea Freight PCBA", total_sea],
        ["Air Freight Cost (USD)", f"${total_air_cost:,.2f}"],
    ]
    for ri, rdata in enumerate(kpi, 1):
        for ci, v in enumerate(rdata, 1):
            c = ws_kpi.cell(ri, ci, v); c.border = bdr
    for c in range(1, 3):
        c2 = ws_kpi.cell(1, c); c2.fill, c2.font = orange, wh_b
    for ri in range(2, len(kpi)+1): ws_kpi.cell(ri, 1).font = bld
    ws_kpi.column_dimensions["A"].width = 25
    ws_kpi.column_dimensions["B"].width = 30

    # ── Sheet 5: Fixture Load Detail ───────────────────────
    ws_fix = wb_out.create_sheet("Fixture_Load_Detail")
    fh = ["Fixture", "Daily Cap"] + [d.strftime("%d/%m") for _, d in date_cols]
    for c, h in enumerate(fh, 1):
        cell = ws_fix.cell(1, c, h)
        cell.fill, cell.font = orange, wh_b
        cell.alignment = Alignment(horizontal="center"); cell.border = bdr
    for fi, (fx, dc) in enumerate(sorted(fixture_daily_cap.items()), 2):
        ws_fix.cell(fi, 1).value = fx; ws_fix.cell(fi, 1).font = bld
        ws_fix.cell(fi, 2).value = round(dc, 1)
        for c in (1, 2): ws_fix.cell(fi, c).border = bdr
        sfx = [s for s in all_skus if sku_fixture.get(s) == fx]
        for t in range(num_days):
            ld = sum(solve_as_int(value(prod[(s, t)])) for s in sfx)
            cl = ws_fix.cell(fi, 3 + t)
            cl.value = ld; cl.border = bdr
            if ld > dc + 0.5:   cl.fill, cl.font = red, bld
            elif ld > dc * 0.9: cl.fill = yellow
    for col_cells in ws_fix.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_fix.column_dimensions[cl].width = min(max(mx + 2, 8), 16)
    ws_fix.freeze_panes = "C2"

    # ── Sheet 6: SKU Inventory Flow ────────────────────────
    ws_iv = wb_out.create_sheet("SKU_Inventory_Flow")
    ivh = ["SKU", "Priority", "Period", "Ship Date", "Beg Inv",
           "Period Prod", "Available", "Demand", "End Inv", "Shortage", "Risk"]
    for c, h in enumerate(ivh, 1):
        cell = ws_iv.cell(1, c, h)
        cell.fill, cell.font = orange, wh_b
        cell.alignment = Alignment(horizontal="center"); cell.border = bdr
    ir = 2
    for s in all_skus:
        pri = priority_map.get(s, 2)
        for j, rd in enumerate(sku_groups[s]):
            i = rd["row"]; ship = rd["ship_request"]
            ps = sku_groups[s][j-1]["ship_request"] if j > 0 else None
            pp2 = sum(solve_as_int(value(prod[(s, t)]))
                      for t, (_, d) in enumerate(date_cols)
                      if (ps is None and d <= ship) or (ps and d > ps and d <= ship))
            bg = INITIAL_FG_INV if j == 0 else solve_as_int(value(inv[sku_groups[s][j-1]["row"]]))
            ei = solve_as_int(value(inv[i]))
            sh = solve_as_int(value(short[i]))
            rk = "SHORTAGE" if sh > 0 else "OK"
            for ci, v in enumerate([s, pri, j+1, ship, bg, pp2, bg+pp2,
                                     rd["demand"], ei, sh, rk], 1):
                cl = ws_iv.cell(ir, ci); cl.value = v; cl.border = bdr
            ws_iv.cell(ir, 4).number_format = dfmt
            rc2 = ws_iv.cell(ir, 11)
            rc2.fill = green if rk == "OK" else red
            if rk != "OK": rc2.font = bld
            ir += 1
    for col_cells in ws_iv.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_iv.column_dimensions[cl].width = min(max(mx + 2, 10), 20)

    # ── Save ────────────────────────────────────────────────
    wb_out.save(OUTPUT_FILE)

    # ── Console ─────────────────────────────────────────────
    print()
    print("=" * 62)
    print("  SCENARIO RESULT: PCBA Firmware Defect (SKU A)")
    print("=" * 62)
    print(f"  Status         : {LpStatus[model.status]}")
    print(f"  Horizon        : {num_days} workdays")
    print(f"  Total demand   : {sum(rd['demand'] for rd in rows):,}")
    print(f"  Total planned  : {tot_plan:,}")
    print(f"  Total shortage : {tot_short:,}")
    print("-" * 62)
    if pri1_dem:
        print(f"  Pri-1 (A,C,E)  : {pri1_short:>8,} / {pri1_dem:>8,}  ({100*pri1_short/pri1_dem:.1f}%)")
    if pri2_dem:
        print(f"  Pri-2 (B,D,F)  : {pri2_short:>8,} / {pri2_dem:>8,}  ({100*pri2_short/pri2_dem:.1f}%)")
    print("-" * 62)
    print(f"  PCBA by AIR    : {total_air:>8,} units")
    print(f"  PCBA by SEA    : {total_sea:>8,} units")
    print(f"  AIR COST       :  ${total_air_cost:>10,.2f}")
    print("-" * 62)
    print(f"  Output: {OUTPUT_FILE}")
    print("=" * 62)


if __name__ == "__main__":
    main()