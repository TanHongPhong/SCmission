
from __future__ import annotations

import math
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import milp, LinearConstraint, Bounds

BASE = Path(__file__).parent
INPUT_XLSX = BASE.parent / "solver.xlsx"

# -----------------------------
# Generic helpers
# -----------------------------

def normalize_market(x):
    if x is None:
        return None
    s = str(x).strip()
    if s.lower() in ["australia", "au"]:
        return "AU"
    return s.upper()

def nice_week_label(x):
    if hasattr(x, "strftime"):
        return x.strftime("%Y-%m-%d")
    return str(x)

def safe_float(x, default=0.0):
    if x in (None, ""):
        return default
    return float(x)

def autosize_and_style(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    border = Border(bottom=Side(style="thin", color="C9C9C9"))
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = border
            if isinstance(c.value, (int, float)):
                header = ws.cell(1, c.column).value or ""
                hs = str(header)
                if "Cost" in hs:
                    c.number_format = '$#,##0.00'
                elif "CBM" in hs:
                    c.number_format = '0.0000'
                elif "Boxes" in hs or hs in ("Week", "n40", "n20", "Load_Index", "Total_n40", "Total_n20", "Markets", "Weeks"):
                    c.number_format = '0'
                else:
                    c.number_format = '0.00'
    max_scan_rows = min(ws.max_row, 120)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for r in range(1, max_scan_rows + 1):
            v = ws.cell(r, col).value
            if v is not None:
                width = max(width, len(str(v)) + 2)
        ws.column_dimensions[letter].width = min(width, 28)

def write_df_sheet(wb, name: str, df: pd.DataFrame):
    ws = wb.create_sheet(name)
    if df is None or df.empty:
        ws["A1"] = "No rows"
        return ws
    ws.append(df.columns.tolist())
    for row in df.itertuples(index=False):
        ws.append(list(row))
    autosize_and_style(ws)
    return ws

# -----------------------------
# Parsing helpers
# -----------------------------

def find_header_row(ws, required_headers, max_rows=200, max_cols=80):
    req = [str(h).strip().lower() for h in required_headers]
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row_vals = [str(ws.cell(r, c).value).strip().lower() if ws.cell(r, c).value is not None else "" for c in range(1, min(ws.max_column, max_cols) + 1)]
        if all(h in row_vals for h in req):
            return r
    return None

def parse_master(ws) -> pd.DataFrame:
    hdr_row = find_header_row(ws, ["Item name", "Market"])
    if hdr_row is None:
        raise ValueError("Could not find 'SKU master' header row with at least 'Item name' and 'Market'.")

    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if v is not None:
            header_map[str(v).strip().lower()] = c

    item_col = header_map.get("item name")
    desc_col = header_map.get("des")
    pack_col = header_map.get("packing size (pcs/carton)", header_map.get("packing size (pcs/ carton)"))
    cbm_col = header_map.get("cbm (100 cartons)", header_map.get(" cbm (100 cartons) "))
    market_col = header_map.get("market")

    if not all([item_col, pack_col, cbm_col, market_col]):
        raise ValueError("SKU master is missing one of: Item name, Packing size (pcs/carton), CBM (100 cartons), Market.")

    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        item = ws.cell(r, item_col).value
        if item in (None, ""):
            if rows:
                break
            continue
        rows.append({
            "Item": str(item).strip(),
            "Description": ws.cell(r, desc_col).value if desc_col else None,
            "PackSize": safe_float(ws.cell(r, pack_col).value),
            "CBM100": safe_float(ws.cell(r, cbm_col).value),
            "CBMPerBox": safe_float(ws.cell(r, cbm_col).value) / 100.0,
            "Market": normalize_market(ws.cell(r, market_col).value),
        })

    if not rows:
        raise ValueError("SKU master header found, but no item rows detected.")
    return pd.DataFrame(rows)

def parse_lanes(ws) -> pd.DataFrame:
    hdr_row = find_header_row(ws, ["Market", "Mode", "Cost", "Cap_CBM"])
    if hdr_row is None:
        raise ValueError("Could not find 'Lane parameters' header row with Market / Mode / Cost / Cap_CBM.")

    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if v is not None:
            header_map[str(v).strip().lower()] = c

    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        market = ws.cell(r, header_map["market"]).value
        mode = ws.cell(r, header_map["mode"]).value
        if market in (None, "") or mode in (None, ""):
            if rows:
                break
            continue
        rows.append({
            "Market": normalize_market(market),
            "Mode": str(mode).strip().replace('"', ''),
            "Cost": safe_float(ws.cell(r, header_map["cost"]).value),
            "Cap_CBM": safe_float(ws.cell(r, header_map["cap_cbm"]).value),
        })

    if not rows:
        raise ValueError("Lane parameter header found, but no lane rows detected.")
    df = pd.DataFrame(rows)
    return df

def parse_decision_table(ws, valid_items: List[str]) -> pd.DataFrame:
    hdr_row = find_header_row(ws, ["Week"], max_rows=250, max_cols=100)
    if hdr_row is None:
        raise ValueError("Could not find Decision table header row containing 'Week'.")

    header_cells = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        header_cells.append(v)

    week_col = None
    item_cols = []
    for idx, val in enumerate(header_cells, start=1):
        if val is None:
            continue
        s = str(val).strip()
        if s.lower() == "week":
            week_col = idx
        elif s in valid_items:
            item_cols.append((idx, s))
        elif s.startswith("Units_") and s.replace("Units_", "") in valid_items:
            item_cols.append((idx, s.replace("Units_", "")))

    if week_col is None or not item_cols:
        raise ValueError("Decision table found, but item columns were not detected next to Week.")

    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        week = ws.cell(r, week_col).value
        if week in (None, ""):
            if rows:
                break
            continue
        row = {"Week": week}
        nonzero = False
        for c, item in item_cols:
            val = ws.cell(r, c).value
            num = safe_float(val, 0.0)
            row[item] = num
            if abs(num) > 1e-12:
                nonzero = True
        rows.append(row)

    if not rows:
        raise ValueError("Decision table header found, but no weekly rows detected.")
    df = pd.DataFrame(rows)
    df.insert(0, "Week_Index", range(1, len(df) + 1))
    return df

# -----------------------------
# Exact global MILP by market-week
# -----------------------------

def solve_exact_market_week(
    week_idx: int,
    week_raw: Any,
    market: str,
    item_demands_units: Dict[str, float],
    master_mkt: pd.DataFrame,
    lane_mkt: pd.DataFrame,
):
    """
    Exact MILP:
    Decide at the same time:
      - y40_j, y20_k: whether each candidate container is used
      - x40_{j,i}, x20_{k,i}: integer number of boxes of item i loaded on each container
      - l_i: integer number of boxes of item i sent by LCL

    Objective:
      min cost40 * sum(y40_j) + cost20 * sum(y20_k) + costLCL * sum(cbm_i * l_i)

    Constraints:
      For each item i:
        sum_j x40_{j,i} + sum_k x20_{k,i} + l_i = demand_boxes_i

      For each 40 container j:
        sum_i cbm_i * x40_{j,i} <= cap40 * y40_j

      For each 20 container k:
        sum_i cbm_i * x20_{k,i} <= cap20 * y20_k

      Symmetry breaking:
        y40_j >= y40_{j+1}
        y20_k >= y20_{k+1}

      All x / l are integer >= 0
      y are binary
    """
    items = master_mkt["Item"].tolist()
    pack = {r["Item"]: float(r["PackSize"]) for _, r in master_mkt.iterrows()}
    cbm = {r["Item"]: float(r["CBMPerBox"]) for _, r in master_mkt.iterrows()}

    boxes_demand = {
        item: int(math.ceil(float(item_demands_units.get(item, 0.0)) / pack[item])) if float(item_demands_units.get(item, 0.0)) > 0 else 0
        for item in items
    }
    total_demand_cbm = sum(boxes_demand[i] * cbm[i] for i in items)

    cost40 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "40", "Cost"].iloc[0])
    cap40 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "40", "Cap_CBM"].iloc[0])
    cost20 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "20", "Cost"].iloc[0])
    cap20 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "20", "Cap_CBM"].iloc[0])
    cost_lcl = float(lane_mkt.loc[lane_mkt["Mode"].astype(str).str.upper() == "LCL", "Cost"].iloc[0])

    if total_demand_cbm <= 1e-12:
        weekly = {
            "Market": market,
            "Week": int(week_idx),
            "Week_Date": nice_week_label(week_raw),
            "Total_Demand_CBM": 0.0,
            "n40": 0,
            "n20": 0,
            "LCL_CBM": 0.0,
            "Coverage_CBM": 0.0,
            "Slack_CBM": 0.0,
            "Cost_40": 0.0,
            "Cost_20": 0.0,
            "Cost_LCL": 0.0,
            "Cost": 0.0,
        }
        for item in items:
            units = float(item_demands_units.get(item, 0.0))
            weekly[f"Units_{item}"] = units
            weekly[f"Boxes_{item}"] = 0 if units <= 0 else int(math.ceil(units / pack[item]))
            weekly[f"CBM_{item}"] = 0.0 if units <= 0 else weekly[f"Boxes_{item}"] * cbm[item]
            weekly[f"LCL_Boxes_{item}"] = 0
            weekly[f"LCL_Units_{item}"] = 0
            weekly[f"LCL_CBM_{item}"] = 0.0
        return weekly, pd.DataFrame()

    # Candidate container counts: enough to cover everything with one type if needed
    J40 = int(math.ceil(total_demand_cbm / cap40))
    J20 = int(math.ceil(total_demand_cbm / cap20))

    # Variable indexing
    var_names = []
    idx = {}

    def add_var(name):
        idx[name] = len(var_names)
        var_names.append(name)

    # x40[j, item]
    for j in range(J40):
        for item in items:
            add_var(f"x40|{j}|{item}")

    # x20[k, item]
    for k in range(J20):
        for item in items:
            add_var(f"x20|{k}|{item}")

    # lcl boxes by item
    for item in items:
        add_var(f"lcl|{item}")

    # y binaries
    for j in range(J40):
        add_var(f"y40|{j}")
    for k in range(J20):
        add_var(f"y20|{k}")

    n = len(var_names)
    c = np.zeros(n)
    lb = np.zeros(n)
    ub = np.full(n, np.inf)
    integrality = np.ones(n)  # all integer; y variables become binary via ub=1

    # Objective and bounds
    for j in range(J40):
        c[idx[f"y40|{j}"]] = cost40
        ub[idx[f"y40|{j}"]] = 1
    for k in range(J20):
        c[idx[f"y20|{k}"]] = cost20
        ub[idx[f"y20|{k}"]] = 1

    for item in items:
        demand_boxes = boxes_demand[item]
        # lcl boxes upper bound
        ub[idx[f"lcl|{item}"]] = demand_boxes
        c[idx[f"lcl|{item}"]] = cost_lcl * cbm[item]

    for j in range(J40):
        for item in items:
            ub[idx[f"x40|{j}|{item}"]] = boxes_demand[item]
    for k in range(J20):
        for item in items:
            ub[idx[f"x20|{k}|{item}"]] = boxes_demand[item]

    # Constraints
    A = []
    lower = []
    upper = []

    # 1) Item conservation
    for item in items:
        row = np.zeros(n)
        for j in range(J40):
            row[idx[f"x40|{j}|{item}"]] = 1
        for k in range(J20):
            row[idx[f"x20|{k}|{item}"]] = 1
        row[idx[f"lcl|{item}"]] = 1
        A.append(row)
        lower.append(boxes_demand[item])
        upper.append(boxes_demand[item])

    # 2) 40-container CBM capacity
    for j in range(J40):
        row = np.zeros(n)
        for item in items:
            row[idx[f"x40|{j}|{item}"]] = cbm[item]
        row[idx[f"y40|{j}"]] = -cap40
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    # 3) 20-container CBM capacity
    for k in range(J20):
        row = np.zeros(n)
        for item in items:
            row[idx[f"x20|{k}|{item}"]] = cbm[item]
        row[idx[f"y20|{k}"]] = -cap20
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    # 4) Symmetry breaking to reduce equivalent solutions
    for j in range(J40 - 1):
        row = np.zeros(n)
        row[idx[f"y40|{j}"]] = -1
        row[idx[f"y40|{j+1}"]] = 1
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    for k in range(J20 - 1):
        row = np.zeros(n)
        row[idx[f"y20|{k}"]] = -1
        row[idx[f"y20|{k+1}"]] = 1
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    constraints = LinearConstraint(np.array(A), np.array(lower), np.array(upper))
    bounds = Bounds(lb, ub)

    res = milp(c=c, constraints=constraints, integrality=integrality, bounds=bounds)

    if not res.success:
        raise RuntimeError(f"MILP failed for {market} week {week_idx}: {res.message}")

    x = np.rint(res.x).astype(int)

    y40 = [int(x[idx[f"y40|{j}"]]) for j in range(J40)]
    y20 = [int(x[idx[f"y20|{k}"]]) for k in range(J20)]

    # Build detail rows
    detail_rows = []
    n40 = sum(y40)
    n20 = sum(y20)
    cost_40 = n40 * cost40
    cost_20 = n20 * cost20

    used_container_cbm = 0.0

    # 40ft detail
    cont_counter = 0
    for j in range(J40):
        if y40[j] != 1:
            continue
        cont_counter += 1
        row = {
            "Market": market,
            "Week": int(week_idx),
            "Week_Date": nice_week_label(week_raw),
            "Load_Type": "40",
            "Load_Index": cont_counter,
            "Load_ID": f"40_{cont_counter}",
            "Capacity_CBM": cap40,
        }
        used_cbm = 0.0
        total_boxes = 0
        total_units = 0.0
        for item in items:
            boxes = int(x[idx[f"x40|{j}|{item}"]])
            units = boxes * pack[item]
            cbm_used = boxes * cbm[item]
            row[f"Boxes_{item}"] = boxes
            row[f"Units_{item}"] = units
            row[f"CBM_{item}"] = round(cbm_used, 6)
            used_cbm += cbm_used
            total_boxes += boxes
            total_units += units
        row["Total_Boxes"] = total_boxes
        row["Total_Units"] = total_units
        row["Used_CBM"] = round(used_cbm, 6)
        row["Slack_CBM"] = round(cap40 - used_cbm, 6)
        used_container_cbm += used_cbm
        detail_rows.append(row)

    # 20ft detail
    cont_counter = 0
    for k in range(J20):
        if y20[k] != 1:
            continue
        cont_counter += 1
        row = {
            "Market": market,
            "Week": int(week_idx),
            "Week_Date": nice_week_label(week_raw),
            "Load_Type": "20",
            "Load_Index": cont_counter,
            "Load_ID": f"20_{cont_counter}",
            "Capacity_CBM": cap20,
        }
        used_cbm = 0.0
        total_boxes = 0
        total_units = 0.0
        for item in items:
            boxes = int(x[idx[f"x20|{k}|{item}"]])
            units = boxes * pack[item]
            cbm_used = boxes * cbm[item]
            row[f"Boxes_{item}"] = boxes
            row[f"Units_{item}"] = units
            row[f"CBM_{item}"] = round(cbm_used, 6)
            used_cbm += cbm_used
            total_boxes += boxes
            total_units += units
        row["Total_Boxes"] = total_boxes
        row["Total_Units"] = total_units
        row["Used_CBM"] = round(used_cbm, 6)
        row["Slack_CBM"] = round(cap20 - used_cbm, 6)
        used_container_cbm += used_cbm
        detail_rows.append(row)

    # LCL detail (one row aggregating all LCL boxes for this week-market)
    lcl_boxes_total = 0
    lcl_units_total = 0.0
    lcl_cbm_total = 0.0
    lcl_row = {
        "Market": market,
        "Week": int(week_idx),
        "Week_Date": nice_week_label(week_raw),
        "Load_Type": "LCL",
        "Load_Index": 1,
        "Load_ID": "LCL",
        "Capacity_CBM": None,
    }
    for item in items:
        boxes = int(x[idx[f"lcl|{item}"]])
        units = boxes * pack[item]
        cbm_used = boxes * cbm[item]
        lcl_row[f"Boxes_{item}"] = boxes
        lcl_row[f"Units_{item}"] = units
        lcl_row[f"CBM_{item}"] = round(cbm_used, 6)
        lcl_boxes_total += boxes
        lcl_units_total += units
        lcl_cbm_total += cbm_used
    lcl_row["Total_Boxes"] = lcl_boxes_total
    lcl_row["Total_Units"] = lcl_units_total
    lcl_row["Used_CBM"] = round(lcl_cbm_total, 6)
    lcl_row["Slack_CBM"] = 0.0
    if lcl_boxes_total > 0:
        detail_rows.append(lcl_row)

    cost_lcl_total = lcl_cbm_total * cost_lcl
    coverage_cbm = n40 * cap40 + n20 * cap20 + lcl_cbm_total
    slack_cbm = coverage_cbm - total_demand_cbm
    total_cost = cost_40 + cost_20 + cost_lcl_total

    weekly = {
        "Market": market,
        "Week": int(week_idx),
        "Week_Date": nice_week_label(week_raw),
    }

    for item in items:
        units = float(item_demands_units.get(item, 0.0))
        bxs = boxes_demand[item]
        cbm_used = bxs * cbm[item]
        lcl_boxes = int(x[idx[f"lcl|{item}"]])
        weekly[f"Units_{item}"] = units
        weekly[f"Boxes_{item}"] = bxs
        weekly[f"CBM_{item}"] = round(cbm_used, 6)
        weekly[f"LCL_Boxes_{item}"] = lcl_boxes
        weekly[f"LCL_Units_{item}"] = lcl_boxes * pack[item]
        weekly[f"LCL_CBM_{item}"] = round(lcl_boxes * cbm[item], 6)

    weekly["Total_Demand_CBM"] = round(total_demand_cbm, 6)
    weekly["n40"] = n40
    weekly["n20"] = n20
    weekly["LCL_CBM"] = round(lcl_cbm_total, 6)
    weekly["Coverage_CBM"] = round(coverage_cbm, 6)
    weekly["Slack_CBM"] = round(slack_cbm, 6)
    weekly["Cost_40"] = round(cost_40, 2)
    weekly["Cost_20"] = round(cost_20, 2)
    weekly["Cost_LCL"] = round(cost_lcl_total, 2)
    weekly["Cost"] = round(total_cost, 2)

    detail_df = pd.DataFrame(detail_rows)
    return weekly, detail_df

# -----------------------------
# Main build
# -----------------------------

def build_outputs(input_path: Path):
    wb = load_workbook(input_path, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Workbook has no sheets.")
    ws = wb[wb.sheetnames[0]]

    master = parse_master(ws)
    lanes = parse_lanes(ws)
    demand = parse_decision_table(ws, master["Item"].tolist())

    weekly_market_rows = []
    detail_by_market = {"US": [], "UK": [], "AU": []}

    for _, week_row in demand.iterrows():
        week_idx = int(week_row["Week_Index"])
        week_raw = week_row["Week"]
        for market in ["US", "UK", "AU"]:
            master_mkt = master[master["Market"] == market].copy()
            if master_mkt.empty:
                continue
            lanes_mkt = lanes[lanes["Market"] == market].copy()
            item_demands_units = {item: float(week_row.get(item, 0.0)) for item in master_mkt["Item"].tolist()}
            weekly, detail_df = solve_exact_market_week(
                week_idx=week_idx,
                week_raw=week_raw,
                market=market,
                item_demands_units=item_demands_units,
                master_mkt=master_mkt,
                lane_mkt=lanes_mkt,
            )
            weekly_market_rows.append(weekly)
            if not detail_df.empty:
                detail_by_market[market].append(detail_df)

    weekly_all = pd.DataFrame(weekly_market_rows)

    outputs = {}
    for market in ["US", "UK", "AU"]:
        outputs[market] = weekly_all[weekly_all["Market"] == market].reset_index(drop=True)
        if detail_by_market[market]:
            outputs[f"{market}_Detail"] = pd.concat(detail_by_market[market], ignore_index=True)
        else:
            outputs[f"{market}_Detail"] = pd.DataFrame()

    all_detail = [outputs["US_Detail"], outputs["UK_Detail"], outputs["AU_Detail"]]
    outputs["All_Load_Detail"] = pd.concat([df for df in all_detail if not df.empty], ignore_index=True) if any(not df.empty for df in all_detail) else pd.DataFrame()

    outputs["Weekly_All_Markets"] = weekly_all.reset_index(drop=True)

    summary_market = (
        weekly_all.groupby("Market", as_index=False)
        .agg(
            Weeks=("Week", "count"),
            Total_Demand_CBM=("Total_Demand_CBM", "sum"),
            Total_n40=("n40", "sum"),
            Total_n20=("n20", "sum"),
            Total_LCL_CBM=("LCL_CBM", "sum"),
            Total_Coverage_CBM=("Coverage_CBM", "sum"),
            Total_Slack_CBM=("Slack_CBM", "sum"),
            Total_Cost_40=("Cost_40", "sum"),
            Total_Cost_20=("Cost_20", "sum"),
            Total_Cost_LCL=("Cost_LCL", "sum"),
            Total_Cost=("Cost", "sum"),
        )
    )
    outputs["Summary_Market"] = summary_market

    summary_all = pd.DataFrame([{
        "Markets": int(summary_market["Market"].nunique()) if not summary_market.empty else 0,
        "Weeks": int(weekly_all["Week"].nunique()) if not weekly_all.empty else 0,
        "Total_Demand_CBM": round(summary_market["Total_Demand_CBM"].sum(), 6) if not summary_market.empty else 0.0,
        "Total_n40": int(summary_market["Total_n40"].sum()) if not summary_market.empty else 0,
        "Total_n20": int(summary_market["Total_n20"].sum()) if not summary_market.empty else 0,
        "Total_LCL_CBM": round(summary_market["Total_LCL_CBM"].sum(), 6) if not summary_market.empty else 0.0,
        "Total_Coverage_CBM": round(summary_market["Total_Coverage_CBM"].sum(), 6) if not summary_market.empty else 0.0,
        "Total_Slack_CBM": round(summary_market["Total_Slack_CBM"].sum(), 6) if not summary_market.empty else 0.0,
        "Total_Cost_40": round(summary_market["Total_Cost_40"].sum(), 2) if not summary_market.empty else 0.0,
        "Total_Cost_20": round(summary_market["Total_Cost_20"].sum(), 2) if not summary_market.empty else 0.0,
        "Total_Cost_LCL": round(summary_market["Total_Cost_LCL"].sum(), 2) if not summary_market.empty else 0.0,
        "Total_Cost": round(summary_market["Total_Cost"].sum(), 2) if not summary_market.empty else 0.0,
    }])
    outputs["Summary_All"] = summary_all

    return outputs

def export_outputs(outputs: Dict[str, pd.DataFrame]):
    out_dir = BASE / "baseline_output"
    out_dir.mkdir(parents=True, exist_ok=True)

    # CSVs -> baseline_output/csv/
    csv_dir = out_dir / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    csv_map = {
        "US": csv_dir / "transport_us_exact_milp.csv",
        "UK": csv_dir / "transport_uk_exact_milp.csv",
        "AU": csv_dir / "transport_au_exact_milp.csv",
        "US_Detail": csv_dir / "transport_us_detail_exact_milp.csv",
        "UK_Detail": csv_dir / "transport_uk_detail_exact_milp.csv",
        "AU_Detail": csv_dir / "transport_au_detail_exact_milp.csv",
        "All_Load_Detail": csv_dir / "transport_all_load_detail_exact_milp.csv",
        "Weekly_All_Markets": csv_dir / "transport_weekly_all_markets_exact_milp.csv",
        "Summary_Market": csv_dir / "transport_summary_market_exact_milp.csv",
        "Summary_All": csv_dir / "transport_summary_all_exact_milp.csv",
    }
    for key, path in csv_map.items():
        outputs[key].to_csv(path, index=False)

    # Excel -> baseline_output/
    out_xlsx = out_dir / "transport_output_baseline.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    info = wb.create_sheet("Run_Info")
    rows = [
        ("Input workbook", INPUT_XLSX.name),
        ("Model type", "Exact global MILP by market-week"),
        ("Decision", "n40, n20, box allocation per container, LCL boxes"),
        ("Objective", "Min Cost40*#40 + Cost20*#20 + CostLCL*LCL_CBM"),
        ("Demand basis", "Demand converted from units to integer boxes by ceil(units / pack size)"),
        ("Capacity basis", "Container limits enforced by CBM, not by box count"),
        ("Integrality", "All box allocations and LCL boxes are integers; no fractional boxes"),
        ("Detail output", "Each used container is listed with exact boxes by SKU; residual boxes go to LCL"),
    ]
    for r, (k, v) in enumerate(rows, start=1):
        info.cell(r, 1).value = k
        info.cell(r, 2).value = v
    for c in info["A"]:
        c.font = Font(bold=True)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 110

    order = ["US", "UK", "AU", "US_Detail", "UK_Detail", "AU_Detail", "All_Load_Detail", "Weekly_All_Markets", "Summary_Market", "Summary_All"]
    for key in order:
        write_df_sheet(wb, key, outputs[key])

    wb.save(out_xlsx)
    return out_xlsx

def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_XLSX}")
    outputs = build_outputs(INPUT_XLSX)
    xlsx_path = export_outputs(outputs)
    print(f"  -> {xlsx_path.relative_to(BASE)}")
    print(f"  -> baseline_output/csv/ (10 files)")

if __name__ == "__main__":
    main()
