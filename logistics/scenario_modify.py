"""
Build FULL scenario logistics — all 6 SKUs (A–F), all markets (US/UK/AU)
========================================================================
1. Reads BOM_demand_Scenario  (columns A–F)  +  FGs & Log information
   from the scenario MPS_MRP workbook.
2. Builds solver_scenario.xlsx in the EXACT same format as solver.xlsx
   so that transport_exact_global_milp.py can consume it directly.
3. Runs the MILP transport solver (reuses build_outputs / export_outputs).
4. Saves ALL outputs (CSVs + Excel) into  logistics/scenario_output/
   with the same file names as the baseline, making comparison trivial.
5. Generates a monthly_logistics_cost.xlsx cost report (scenario vs baseline).
"""

from __future__ import annotations

import math
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ───────────────────────────────────────────────────
BASE         = Path(__file__).parent
MPS_MRP      = BASE / "SCM_round2.1_scenario_MPS_MRP.xlsx"
ORIG_SOLVER  = BASE / "solver.xlsx"          # baseline, for reference
OUT_DIR      = BASE / "scenario_modify_output" # ← separate output folder for modified scenario
SCENARIO_SOLVER = OUT_DIR / "solver_scenario_modify.xlsx"
MONTHLY_OUT     = OUT_DIR / "monthly_logistics_cost.xlsx"

# ── Styles ──────────────────────────────────────────────────
NAVY   = PatternFill("solid", fgColor="1F4E78")
ORANGE = PatternFill("solid", fgColor="E65100")
TEAL   = PatternFill("solid", fgColor="00897B")
GREEN  = PatternFill("solid", fgColor="E2F0D9")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED    = PatternFill("solid", fgColor="FDE9D9")
WH_B   = Font(color="FFFFFF", bold=True)
BLD    = Font(bold=True)
THIN   = Side(style="thin", color="BFBFBF")
BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALL_SKUS   = ["A", "B", "C", "D", "E", "F"]
ALL_MARKETS = ["US", "UK", "AU"]


# ─────────────────────────────────────────────────────────────
#  STEP 1 — Build solver_scenario.xlsx  (same format as solver.xlsx)
# ─────────────────────────────────────────────────────────────

def styled_hdr(ws, row, headers, fill=ORANGE):
    for c, h in enumerate(headers, 1):
        cl = ws.cell(row, c, h)
        cl.fill, cl.font = fill, WH_B
        cl.alignment = Alignment(horizontal="center", wrap_text=True)
        cl.border = BDR


def auto_width(ws, max_w=22):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[letter].width = min(max(w + 3, 10), max_w)


def build_scenario_solver_xlsx() -> Path:
    """
    Build solver_scenario.xlsx that transport_exact_global_milp.py can read.
    Exact same layout as solver.xlsx:
      Section 1 – SKU Master  (Item name / Des / Packing size / CBM / Price / Market)
      Section 2 – Lane Parameters (Market / Mode / Cost / Cap_CBM)
      Section 3 – Decision Table  (Week / A / B / C / D / E / F)
    """
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("[1/4] Reading BOM_demand_Scenario  +  FGs & Log information …")

    # ── Read FGs & Log information (SKU master) ─────────────
    fgs_df = pd.read_excel(MPS_MRP, sheet_name="FGs & Log information ", header=None)
    sku_master: list[dict] = []
    for r in range(2, fgs_df.shape[0]):
        item = fgs_df.iloc[r, 1]
        if pd.isna(item) or str(item).strip() == "":
            if sku_master:
                break
            continue
        item = str(item).strip()
        if item not in ALL_SKUS:
            continue
        desc      = str(fgs_df.iloc[r, 2])
        pack_size = int(float(fgs_df.iloc[r, 3]))
        cbm100    = float(fgs_df.iloc[r, 4])
        price     = float(fgs_df.iloc[r, 5])
        market    = str(fgs_df.iloc[r, 6]).strip()
        sku_master.append({
            "Item name": item,
            "Des": desc,
            "Packing size (pcs/carton)": pack_size,
            "CBM (100 cartons)": cbm100,
            "Ex. Work price (USD/pc)": price,
            "Market": market,
        })
    print(f"    SKU master: {[s['Item name'] for s in sku_master]}")

    # ── Read BOM_demand_Scenario (weekly demand, cols A–F) ──
    bom_df = pd.read_excel(MPS_MRP, sheet_name="BOM_demand_Scenario", header=None)
    header = [str(bom_df.iloc[0, c]).strip() for c in range(bom_df.shape[1])]
    week_col = header.index("Week")

    # find column index for each SKU present in the header
    sku_cols: dict[str, int] = {}
    for sku in ALL_SKUS:
        if sku in header:
            sku_cols[sku] = header.index(sku)

    weekly_rows: list[dict] = []
    for r in range(1, bom_df.shape[0]):
        wk = bom_df.iloc[r, week_col]
        if pd.isna(wk):
            break
        row: dict = {"Week": pd.Timestamp(wk)}
        for sku, col in sku_cols.items():
            val = bom_df.iloc[r, col]
            row[sku] = int(float(val)) if not pd.isna(val) else 0
        weekly_rows.append(row)

    print(f"    Found {len(weekly_rows)} weekly rows  "
          f"({weekly_rows[0]['Week'].date()} -> {weekly_rows[-1]['Week'].date()})")

    # ── Lane parameters (identical to solver.xlsx) ──────────
    lane_params = [
        {"Market": "US", "Mode": "40",  "Cost": 5200, "Cap_CBM": 65.0},
        {"Market": "US", "Mode": "20",  "Cost": 3000, "Cap_CBM": 28.0},
        {"Market": "US", "Mode": "LCL", "Cost": 200,  "Cap_CBM": 999999},
        {"Market": "UK", "Mode": "40",  "Cost": 4200, "Cap_CBM": 65.0},
        {"Market": "UK", "Mode": "20",  "Cost": 2500, "Cap_CBM": 28.0},
        {"Market": "UK", "Mode": "LCL", "Cost": 70,   "Cap_CBM": 999999},
        {"Market": "AU", "Mode": "40",  "Cost": 2000, "Cap_CBM": 65.0},
        {"Market": "AU", "Mode": "20",  "Cost": 1100, "Cap_CBM": 28.0},
        {"Market": "AU", "Mode": "LCL", "Cost": 35,   "Cap_CBM": 999999},
    ]

    # ── Write solver_scenario.xlsx ──────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "solver"     # same sheet name as baseline

    r = 1
    # Title rows (like baseline)
    ws.cell(r, 1, "Transport Cost Optimization Model (SCENARIO)").font = BLD
    r += 1
    ws.cell(r, 1, "Solver-ready layout using weekly demand from BOM_demand_Scenario").font = Font(italic=True)
    r += 2  # blank row

    # Section 1 — SKU Master
    ws.cell(r, 1, "SKU master").font = BLD
    r += 1
    master_hdrs = ["Item name", "Des", "Packing size (pcs/carton)",
                   "CBM (100 cartons)", "Ex. Work price (USD/pc)", "Market"]
    for c, h in enumerate(master_hdrs, 1):
        ws.cell(r, c, h).font = BLD
    r += 1
    for sm in sku_master:
        for c, k in enumerate(master_hdrs, 1):
            ws.cell(r, c, sm[k])
        r += 1

    r += 1  # blank
    # Section 2 — Lane Parameters
    ws.cell(r, 1, "Lane parameters").font = BLD
    r += 1
    lane_hdrs = ["Market", "Mode", "Cost", "Cap_CBM"]
    for c, h in enumerate(lane_hdrs, 1):
        ws.cell(r, c, h).font = BLD
    r += 1
    for lp in lane_params:
        for c, k in enumerate(lane_hdrs, 1):
            ws.cell(r, c, lp[k])
        r += 1

    r += 2  # blank rows (like baseline)

    # Section 3 — Decision Table (Week, A, B, C, D, E, F)
    ws.cell(r, 1, "Week").font = BLD
    for ci, sku in enumerate(ALL_SKUS, 2):
        ws.cell(r, ci, sku).font = BLD
    r += 1
    for wr in weekly_rows:
        ws.cell(r, 1, wr["Week"])
        ws.cell(r, 1).number_format = "YYYY-MM-DD"
        for ci, sku in enumerate(ALL_SKUS, 2):
            ws.cell(r, ci, wr.get(sku, 0))
        r += 1

    auto_width(ws)
    wb.save(SCENARIO_SOLVER)
    print(f"    Saved: {SCENARIO_SOLVER.name}")
    return SCENARIO_SOLVER


# ─────────────────────────────────────────────────────────────
#  STEP 2 — Custom MILP function with market-based weights
# ─────────────────────────────────────────────────────────────

def custom_solve_exact_market_week(
    week_idx: int,
    week_raw: any,
    market: str,
    item_demands_units: dict,
    master_mkt: pd.DataFrame,
    lane_mkt: pd.DataFrame,
):
    import numpy as np
    from scipy.optimize import milp, LinearConstraint, Bounds
    import transport_exact_global_milp as milp_mod

    items = master_mkt["Item"].tolist()
    pack = {r["Item"]: float(r["PackSize"]) for _, r in master_mkt.iterrows()}
    cbm = {r["Item"]: float(r["CBMPerBox"]) for _, r in master_mkt.iterrows()}

    boxes_demand = {
        item: int(math.ceil(float(item_demands_units.get(item, 0.0)) / pack[item])) if float(item_demands_units.get(item, 0.0)) > 0 else 0
        for item in items
    }
    total_demand_cbm = sum(boxes_demand[i] * cbm[i] for i in items)

    cost40 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "40", "Cost"].iloc[0])
    cap40 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "40", "Cap_CBM"].iloc[0])
    cost20 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "20", "Cost"].iloc[0])
    cap20 = float(lane_mkt.loc[lane_mkt["Mode"].astype(str) == "20", "Cap_CBM"].iloc[0])
    cost_lcl = float(lane_mkt.loc[lane_mkt["Mode"].astype(str).str.upper() == "LCL", "Cost"].iloc[0])

    if total_demand_cbm <= 1e-12:
        weekly = {
            "Market": market,
            "Week": int(week_idx),
            "Week_Date": milp_mod.nice_week_label(week_raw),
            "Total_Demand_CBM": 0.0,
            "n40": 0,
            "n20": 0,
            "LCL_CBM": 0.0,
            "Coverage_CBM": 0.0,
            "Slack_CBM": 0.0,
            "Cost_40": 0.0,
            "Cost_20": 0.0,
            "Cost_LCL": 0.0,
            "Cost": 0.0,
        }
        for item in items:
            units = float(item_demands_units.get(item, 0.0))
            weekly[f"Units_{item}"] = units
            weekly[f"Boxes_{item}"] = 0 if units <= 0 else int(math.ceil(units / pack[item]))
            weekly[f"CBM_{item}"] = 0.0 if units <= 0 else weekly[f"Boxes_{item}"] * cbm[item]
            weekly[f"LCL_Boxes_{item}"] = 0
            weekly[f"LCL_Units_{item}"] = 0
            weekly[f"LCL_CBM_{item}"] = 0.0
        return weekly, pd.DataFrame()

    J40 = int(math.ceil(total_demand_cbm / cap40))
    J20 = int(math.ceil(total_demand_cbm / cap20))

    var_names = []
    idx = {}

    def add_var(name):
        idx[name] = len(var_names)
        var_names.append(name)

    for j in range(J40):
        for item in items:
            add_var(f"x40|{j}|{item}")

    for k in range(J20):
        for item in items:
            add_var(f"x20|{k}|{item}")

    for item in items:
        add_var(f"lcl|{item}")

    for j in range(J40):
        add_var(f"y40|{j}")
    for k in range(J20):
        add_var(f"y20|{k}")

    n = len(var_names)
    c = np.zeros(n)
    lb = np.zeros(n)
    ub = np.full(n, np.inf)
    integrality = np.ones(n)

    # --- WEIGHTS LOGIC ---
    if market == "US":
        m2 = 0.6
    elif market == "UK":
        m2 = 0.7
    elif market == "AU":
        m2 = 0.7
    else:
        m2 = 0.5
        
    m1 = 1.0 - m2

    for j in range(J40):
        c[idx[f"y40|{j}"]] = cost40 * m1
        ub[idx[f"y40|{j}"]] = 1
    for k in range(J20):
        c[idx[f"y20|{k}"]] = cost20 * m1
        ub[idx[f"y20|{k}"]] = 1

    for item in items:
        demand_boxes = boxes_demand[item]
        ub[idx[f"lcl|{item}"]] = demand_boxes
        c[idx[f"lcl|{item}"]] = cost_lcl * cbm[item] * m2

    for j in range(J40):
        for item in items:
            ub[idx[f"x40|{j}|{item}"]] = boxes_demand[item]
    for k in range(J20):
        for item in items:
            ub[idx[f"x20|{k}|{item}"]] = boxes_demand[item]

    A = []
    lower = []
    upper = []

    for item in items:
        row = np.zeros(n)
        for j in range(J40):
            row[idx[f"x40|{j}|{item}"]] = 1
        for k in range(J20):
            row[idx[f"x20|{k}|{item}"]] = 1
        row[idx[f"lcl|{item}"]] = 1
        A.append(row)
        lower.append(boxes_demand[item])
        upper.append(boxes_demand[item])

    for j in range(J40):
        row = np.zeros(n)
        for item in items:
            row[idx[f"x40|{j}|{item}"]] = cbm[item]
        row[idx[f"y40|{j}"]] = -cap40
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    for k in range(J20):
        row = np.zeros(n)
        for item in items:
            row[idx[f"x20|{k}|{item}"]] = cbm[item]
        row[idx[f"y20|{k}"]] = -cap20
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    for j in range(J40 - 1):
        row = np.zeros(n)
        row[idx[f"y40|{j}"]] = -1
        row[idx[f"y40|{j+1}"]] = 1
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    for k in range(J20 - 1):
        row = np.zeros(n)
        row[idx[f"y20|{k}"]] = -1
        row[idx[f"y20|{k+1}"]] = 1
        A.append(row)
        lower.append(-np.inf)
        upper.append(0.0)

    constraints = LinearConstraint(np.array(A), np.array(lower), np.array(upper))
    bounds = Bounds(lb, ub)

    res = milp(c=c, constraints=constraints, integrality=integrality, bounds=bounds)

    if not res.success:
        raise RuntimeError(f"MILP failed for {market} week {week_idx}: {res.message}")

    x = np.rint(res.x).astype(int)

    y40 = [int(x[idx[f"y40|{j}"]]) for j in range(J40)]
    y20 = [int(x[idx[f"y20|{k}"]]) for k in range(J20)]

    detail_rows = []
    n40 = sum(y40)
    n20 = sum(y20)
    cost_40 = n40 * cost40
    cost_20 = n20 * cost20

    used_container_cbm = 0.0

    cont_counter = 0
    for j in range(J40):
        if y40[j] != 1:
            continue
        cont_counter += 1
        row = {
            "Market": market,
            "Week": int(week_idx),
            "Week_Date": milp_mod.nice_week_label(week_raw),
            "Load_Type": "40",
            "Load_Index": cont_counter,
            "Load_ID": f"40_{cont_counter}",
            "Capacity_CBM": cap40,
        }
        used_cbm = 0.0
        total_boxes = 0
        total_units = 0.0
        for item in items:
            boxes = int(x[idx[f"x40|{j}|{item}"]])
            units = boxes * pack[item]
            cbm_used = boxes * cbm[item]
            row[f"Boxes_{item}"] = boxes
            row[f"Units_{item}"] = units
            row[f"CBM_{item}"] = round(cbm_used, 6)
            used_cbm += cbm_used
            total_boxes += boxes
            total_units += units
        row["Total_Boxes"] = total_boxes
        row["Total_Units"] = total_units
        row["Used_CBM"] = round(used_cbm, 6)
        row["Slack_CBM"] = round(cap40 - used_cbm, 6)
        used_container_cbm += used_cbm
        detail_rows.append(row)

    cont_counter = 0
    for k in range(J20):
        if y20[k] != 1:
            continue
        cont_counter += 1
        row = {
            "Market": market,
            "Week": int(week_idx),
            "Week_Date": milp_mod.nice_week_label(week_raw),
            "Load_Type": "20",
            "Load_Index": cont_counter,
            "Load_ID": f"20_{cont_counter}",
            "Capacity_CBM": cap20,
        }
        used_cbm = 0.0
        total_boxes = 0
        total_units = 0.0
        for item in items:
            boxes = int(x[idx[f"x20|{k}|{item}"]])
            units = boxes * pack[item]
            cbm_used = boxes * cbm[item]
            row[f"Boxes_{item}"] = boxes
            row[f"Units_{item}"] = units
            row[f"CBM_{item}"] = round(cbm_used, 6)
            used_cbm += cbm_used
            total_boxes += boxes
            total_units += units
        row["Total_Boxes"] = total_boxes
        row["Total_Units"] = total_units
        row["Used_CBM"] = round(used_cbm, 6)
        row["Slack_CBM"] = round(cap20 - used_cbm, 6)
        used_container_cbm += used_cbm
        detail_rows.append(row)

    lcl_boxes_total = 0
    lcl_units_total = 0.0
    lcl_cbm_total = 0.0
    lcl_row = {
        "Market": market,
        "Week": int(week_idx),
        "Week_Date": milp_mod.nice_week_label(week_raw),
        "Load_Type": "LCL",
        "Load_Index": 1,
        "Load_ID": "LCL",
        "Capacity_CBM": None,
    }
    for item in items:
        boxes = int(x[idx[f"lcl|{item}"]])
        units = boxes * pack[item]
        cbm_used = boxes * cbm[item]
        lcl_row[f"Boxes_{item}"] = boxes
        lcl_row[f"Units_{item}"] = units
        lcl_row[f"CBM_{item}"] = round(cbm_used, 6)
        lcl_boxes_total += boxes
        lcl_units_total += units
        lcl_cbm_total += cbm_used
    lcl_row["Total_Boxes"] = lcl_boxes_total
    lcl_row["Total_Units"] = lcl_units_total
    lcl_row["Used_CBM"] = round(lcl_cbm_total, 6)
    lcl_row["Slack_CBM"] = 0.0
    if lcl_boxes_total > 0:
        detail_rows.append(lcl_row)

    cost_lcl_total = lcl_cbm_total * cost_lcl
    coverage_cbm = n40 * cap40 + n20 * cap20 + lcl_cbm_total
    slack_cbm = coverage_cbm - total_demand_cbm
    total_cost = cost_40 + cost_20 + cost_lcl_total

    weekly = {
        "Market": market,
        "Week": int(week_idx),
        "Week_Date": milp_mod.nice_week_label(week_raw),
    }

    for item in items:
        units = float(item_demands_units.get(item, 0.0))
        bxs = boxes_demand[item]
        cbm_used = bxs * cbm[item]
        lcl_boxes = int(x[idx[f"lcl|{item}"]])
        weekly[f"Units_{item}"] = units
        weekly[f"Boxes_{item}"] = bxs
        weekly[f"CBM_{item}"] = round(cbm_used, 6)
        weekly[f"LCL_Boxes_{item}"] = lcl_boxes
        weekly[f"LCL_Units_{item}"] = lcl_boxes * pack[item]
        weekly[f"LCL_CBM_{item}"] = round(lcl_boxes * cbm[item], 6)

    weekly["Total_Demand_CBM"] = round(total_demand_cbm, 6)
    weekly["n40"] = n40
    weekly["n20"] = n20
    weekly["LCL_CBM"] = round(lcl_cbm_total, 6)
    weekly["Coverage_CBM"] = round(coverage_cbm, 6)
    weekly["Slack_CBM"] = round(slack_cbm, 6)
    weekly["Cost_40"] = round(cost_40, 2)
    weekly["Cost_20"] = round(cost_20, 2)
    weekly["Cost_LCL"] = round(cost_lcl_total, 2)
    weekly["Cost"] = round(total_cost, 2)

    detail_df = pd.DataFrame(detail_rows)
    return weekly, detail_df


# ─────────────────────────────────────────────────────────────
#  STEP 3 — Run MILP solver  +  Export (reuse baseline code)
# ─────────────────────────────────────────────────────────────

def run_milp_and_export(solver_path: Path):
    """
    Re-use the exact same build_outputs() + export_outputs() logic
    from transport_exact_global_milp.py, but pointed at our scenario
    solver file and writing into scenario_output/.
    """
    print("[2/3] Running MILP transport solver ...")

    sys.path.insert(0, str(BASE))
    import transport_exact_global_milp as milp_mod

    # Patch the exact solver to use the new weights logic
    milp_mod.solve_exact_market_week = custom_solve_exact_market_week

    # Build the outputs dict (identical structure to baseline)
    outputs = milp_mod.build_outputs(solver_path)

    # Export
    print("[3/3] Exporting outputs …")

    # CSVs -> scenario_output/csv/
    csv_dir = OUT_DIR / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    csv_map = {
        "US":                csv_dir / "transport_us_exact_milp.csv",
        "UK":                csv_dir / "transport_uk_exact_milp.csv",
        "AU":                csv_dir / "transport_au_exact_milp.csv",
        "US_Detail":         csv_dir / "transport_us_detail_exact_milp.csv",
        "UK_Detail":         csv_dir / "transport_uk_detail_exact_milp.csv",
        "AU_Detail":         csv_dir / "transport_au_detail_exact_milp.csv",
        "All_Load_Detail":   csv_dir / "transport_all_load_detail_exact_milp.csv",
        "Weekly_All_Markets":csv_dir / "transport_weekly_all_markets_exact_milp.csv",
        "Summary_Market":    csv_dir / "transport_summary_market_exact_milp.csv",
        "Summary_All":       csv_dir / "transport_summary_all_exact_milp.csv",
    }
    for key, path in csv_map.items():
        outputs[key].to_csv(path, index=False)

    # Excel workbook -> scenario_output/  (top-level)
    out_xlsx = OUT_DIR / "transport_output_scenario.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    info = wb.create_sheet("Run_Info")
    rows = [
        ("Input workbook", solver_path.name),
        ("Model type", "Exact global MILP by market-week (SCENARIO)"),
        ("Decision", "n40, n20, box allocation per container, LCL boxes"),
        ("Objective", "Min Cost40*#40 + Cost20*#20 + CostLCL*LCL_CBM"),
        ("Demand basis", "Demand converted from units to integer boxes by ceil(units / pack size)"),
        ("Capacity basis", "Container limits enforced by CBM, not by box count"),
        ("Integrality", "All box allocations and LCL boxes are integers; no fractional boxes"),
        ("Detail output", "Each used container is listed with exact boxes by SKU; residual boxes go to LCL"),
    ]
    for ri, (k, v) in enumerate(rows, start=1):
        info.cell(ri, 1).value = k
        info.cell(ri, 2).value = v
    for c in info["A"]:
        c.font = Font(bold=True)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 110

    order = ["US", "UK", "AU",
             "US_Detail", "UK_Detail", "AU_Detail",
             "All_Load_Detail", "Weekly_All_Markets",
             "Summary_Market", "Summary_All"]
    for key in order:
        milp_mod.write_df_sheet(wb, key, outputs[key])

    wb.save(out_xlsx)
    print(f"    -> {out_xlsx.name}")
    print(f"    -> csv/ ({len(csv_map)} files)")

    return outputs


# ─────────────────────────────────────────────────────────────
#  STEP 4 — Monthly cost report
# ─────────────────────────────────────────────────────────────

def build_monthly_cost_report(outputs: dict):
    """Build monthly logistics cost report with scenario vs baseline."""
    print("    -> monthly_logistics_cost.xlsx")

    weekly_all = outputs["Weekly_All_Markets"].copy()

    # Ensure Week_Date is usable
    def get_month(x):
        if hasattr(x, "month"):
            return x.strftime("%Y-%m")
        try:
            return pd.to_datetime(x).strftime("%Y-%m")
        except Exception:
            return "Unknown"

    weekly_all["Month"] = weekly_all["Week_Date"].apply(get_month)

    # Monthly summary
    monthly = (
        weekly_all.groupby(["Month", "Market"], as_index=False)
        .agg(
            Weeks=("Week", "count"),
            Total_CBM=("Total_Demand_CBM", "sum"),
            n40=("n40", "sum"),
            n20=("n20", "sum"),
            LCL_CBM=("LCL_CBM", "sum"),
            Cost_40=("Cost_40", "sum"),
            Cost_20=("Cost_20", "sum"),
            Cost_LCL=("Cost_LCL", "sum"),
            Total_Cost=("Cost", "sum"),
        )
    ).sort_values(["Month", "Market"])

    wb = Workbook()
    # ── Sheet 1: Monthly by Market ──────────────────────────
    ws = wb.active
    ws.title = "Monthly_Logistics_Cost"
    styled_hdr(ws, 1, [
        "Month", "Market", "Weeks", "Total CBM",
        "40ft Ctnr", "20ft Ctnr", "LCL CBM",
        "Cost 40ft", "Cost 20ft", "Cost LCL", "Total Cost"
    ], TEAL)

    for ri, row in enumerate(monthly.itertuples(index=False), 2):
        vals = [row.Month, row.Market, row.Weeks, round(row.Total_CBM, 2),
                row.n40, row.n20, round(row.LCL_CBM, 2),
                round(row.Cost_40, 2), round(row.Cost_20, 2),
                round(row.Cost_LCL, 2), round(row.Total_Cost, 2)]
        for ci, v in enumerate(vals, 1):
            cl = ws.cell(ri, ci, v)
            cl.border = BDR
            if ci >= 8:
                cl.number_format = "$#,##0.00"
            elif ci == 4 or ci == 7:
                cl.number_format = "0.00"

    # Grand total row
    ri_total = len(monthly) + 2
    ws.cell(ri_total, 1, "GRAND TOTAL").font = WH_B
    ws.cell(ri_total, 1).fill = ORANGE
    for ci, val in enumerate([
        "", "", monthly["Weeks"].sum(), round(monthly["Total_CBM"].sum(), 2),
        monthly["n40"].sum(), monthly["n20"].sum(), round(monthly["LCL_CBM"].sum(), 2),
        round(monthly["Cost_40"].sum(), 2), round(monthly["Cost_20"].sum(), 2),
        round(monthly["Cost_LCL"].sum(), 2), round(monthly["Total_Cost"].sum(), 2)
    ], 2):
        cl = ws.cell(ri_total, ci, val)
        cl.font = BLD
        cl.border = BDR
        if ci >= 8:
            cl.number_format = "$#,##0.00"
    auto_width(ws)
    ws.freeze_panes = "A2"

    # ── Sheet 2: Weekly detail ──────────────────────────────
    ws2 = wb.create_sheet("Weekly_Detail")
    cols_to_show = ["Month", "Market", "Week", "Week_Date",
                    "Total_Demand_CBM", "n40", "n20", "LCL_CBM",
                    "Cost_40", "Cost_20", "Cost_LCL", "Cost"]
    weekly_display = weekly_all[cols_to_show].copy()
    styled_hdr(ws2, 1, weekly_display.columns.tolist(), NAVY)
    for ri, row in enumerate(weekly_display.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            cl = ws2.cell(ri, ci, v)
            cl.border = BDR
            hdr = weekly_display.columns[ci - 1]
            if "Cost" in hdr:
                cl.number_format = "$#,##0.00"
            elif "CBM" in hdr:
                cl.number_format = "0.0000"
    auto_width(ws2)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Summary by Market ──────────────────────────
    ws3 = wb.create_sheet("Summary_by_Market")
    mkt_summary = monthly.groupby("Market", as_index=False).agg(
        Total_CBM=("Total_CBM", "sum"),
        Total_n40=("n40", "sum"),
        Total_n20=("n20", "sum"),
        Total_LCL_CBM=("LCL_CBM", "sum"),
        Total_Cost_40=("Cost_40", "sum"),
        Total_Cost_20=("Cost_20", "sum"),
        Total_Cost_LCL=("Cost_LCL", "sum"),
        Total_Cost=("Total_Cost", "sum"),
    )
    styled_hdr(ws3, 1, mkt_summary.columns.tolist(), ORANGE)
    for ri, row in enumerate(mkt_summary.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            cl = ws3.cell(ri, ci, v)
            cl.border = BDR
            if "Cost" in mkt_summary.columns[ci - 1]:
                cl.number_format = "$#,##0.00"
    auto_width(ws3)

    # ── Sheet 4: Scenario Cost Breakdown ────────────────────
    ws4 = wb.create_sheet("Scenario_Cost_Breakdown")
    styled_hdr(ws4, 1, ["Cost Category", "Amount (USD)", "Notes"], TEAL)

    ocean_cost = round(monthly["Total_Cost"].sum(), 2)
    pcba_air   = 70_040.00      # from scenario analysis (PCBA air freight for A)

    us_cost  = round(monthly.loc[monthly["Market"] == "US", "Total_Cost"].sum(), 2)
    uk_cost  = round(monthly.loc[monthly["Market"] == "UK", "Total_Cost"].sum(), 2)
    au_cost  = round(monthly.loc[monthly["Market"] == "AU", "Total_Cost"].sum(), 2)

    total_scen = ocean_cost + pcba_air

    cost_breakdown = [
        ["--- Ocean Freight (FG Shipment) ---", "", ""],
        ["  US market (A+B+F)", us_cost, "Containers + LCL"],
        ["  UK market (C+D)", uk_cost, "Containers + LCL"],
        ["  AU market (E)", au_cost, "Containers + LCL"],
        ["  Total Ocean Freight", ocean_cost, "Sum of all markets"],
        ["", "", ""],
        ["--- PCBA Inbound Air Freight (Scenario) ---", "", ""],
        ["  PCBA units shipped by air", 35_020, "= 60% recovery"],
        ["  Air premium per PCBA", 2.0, "$2.00 / unit"],
        ["  Total PCBA Air Freight", pcba_air, "= 35,020 × $2"],
        ["", "", ""],
        ["--- Total Scenario Logistics Cost ---", "", ""],
        ["  Ocean Freight", ocean_cost, "FG delivery to customers"],
        ["  PCBA Air Freight", pcba_air, "Component inbound (scenario only)"],
        ["  GRAND TOTAL", total_scen, "Full scenario logistics cost"],
        ["", "", ""],
        ["--- Baseline Comparison ---", "", ""],
        ["  Baseline ocean freight (from solver.xlsx)", "", "Run baseline MILP separately"],
        ["  Additional PCBA air cost", pcba_air, "= Scenario-only cost"],
    ]

    for ri, row in enumerate(cost_breakdown, 2):
        bold = row[0].startswith("---") or "GRAND" in row[0] or "Additional" in row[0]
        for ci, v in enumerate(row, 1):
            cl = ws4.cell(ri, ci, v)
            cl.border = BDR
            if bold:
                cl.font = BLD
                cl.fill = YELLOW
            if ci == 2 and isinstance(v, (int, float)) and v != 0:
                cl.number_format = "$#,##0.00"
    auto_width(ws4, 45)

    wb.save(MONTHLY_OUT)
    print(f"    Saved: {MONTHLY_OUT.name}")
    return monthly


# ─────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  SCENARIO LOGISTICS  (A-F, US/UK/AU)")
    print("=" * 60)

    solver_path = build_scenario_solver_xlsx()
    outputs = run_milp_and_export(solver_path)
    build_monthly_cost_report(outputs)

    # ── Console summary ─────────────────────────────────────
    summary = outputs["Summary_Market"]
    print()
    for _, row in summary.iterrows():
        print("  {:>2}: CBM={:>8.1f}  40ft={:>3}  20ft={:>3}  LCL={:>7.2f}  Cost=${:>10,.2f}".format(
            row["Market"], row["Total_Demand_CBM"],
            int(row["Total_n40"]), int(row["Total_n20"]),
            row["Total_LCL_CBM"], row["Total_Cost"]))
    total_ocean = summary["Total_Cost"].sum()
    print(f"  {'':->58}")
    print("  Total ocean : ${:>10,.2f}".format(total_ocean))
    print("  PCBA air    : $  70,040.00")
    print("  GRAND TOTAL : ${:>10,.2f}".format(total_ocean + 70_040))
    print("=" * 60)


if __name__ == "__main__":
    main()
