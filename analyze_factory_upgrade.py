"""Analyze MPS shortage and propose factory upgrade plan."""
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

f = r'D:\A UEH_UNIVERSITY\MONEY\SCmission\SCM_round2.1_final.xlsx'
wb = load_workbook(f, data_only=True)

ws = wb['MPS']
hdrs = {}
for c in range(1, ws.max_column+1):
    v = str(ws.cell(1,c).value or '').strip()
    hdrs[v] = c

fix_week = defaultdict(lambda: {'dem':0,'prod':0,'cap':0,'short':0,'skus':[]})
sku_shortage = defaultdict(int)
sku_dem_total = defaultdict(int)

for r in range(2, ws.max_row+1):
    sku = ws.cell(r,1).value
    if sku is None: break
    s   = str(sku).strip()
    fix = str(ws.cell(r,2).value or '').strip()
    wk  = str(ws.cell(r,3).value)[:10]
    cap = int(float(ws.cell(r,4).value or 0))
    dem = int(float(ws.cell(r,5).value or 0))
    prod= int(float(ws.cell(r,10).value or 0))
    short=int(float(ws.cell(r,12).value or 0))
    key = (fix, wk)
    fix_week[key]['dem']  += dem
    fix_week[key]['prod'] += prod
    fix_week[key]['short']+= short
    fix_week[key]['cap']   = cap
    if short > 0: fix_week[key]['skus'].append(s)
    sku_shortage[s] += short
    sku_dem_total[s] += dem

# FGs prices
ws_fgs = wb['FGs & Log information ']
sku_price = {}
for r in range(3, 12):
    sku = ws_fgs.cell(r,2).value
    price = ws_fgs.cell(r,6).value
    if sku and price:
        sku_price[str(sku).strip()] = float(price)

print("="*70)
print("  FACTORY BOTTLENECK & UPGRADE ANALYSIS")
print("="*70)

# 1. Shortage summary
print("\n[1] SHORTAGE SUMMARY (MPS Baseline)")
print("{:>4} {:>10} {:>10} {:>10} {:>8}  Price".format(
    "SKU","TotalDem","Shortage","Short%","RevLost"))
print("-"*65)
total_rev = 0
for s in sorted(sku_shortage):
    dem   = sku_dem_total[s]
    short = sku_shortage[s]
    pct   = 100*short/dem if dem > 0 else 0
    price = sku_price.get(s, 0)
    rev   = short * price
    total_rev += rev
    print("{:>4} {:>10,} {:>10,} {:>9.1f}% {:>10,.0f}  {:.2f}".format(
        s, dem, short, pct, rev, price))
print("-"*65)
print("{:>4} {:>10,} {:>10,}            {:>10,.0f}".format(
    "TOT", sum(sku_dem_total.values()), sum(sku_shortage.values()), total_rev))

# 2. Fixture bottleneck
print("\n[2] FIXTURE BOTTLENECK ANALYSIS")
fix_totals = defaultdict(lambda: {'dem':0,'short':0,'cap':0,'weeks_100pct':0,'peak_short':0,'peak_wk':''})
for (fix, wk), v in fix_week.items():
    fix_totals[fix]['dem']   += v['dem']
    fix_totals[fix]['short'] += v['short']
    fix_totals[fix]['cap']    = v['cap']
    if v['short'] > 0:
        fix_totals[fix]['weeks_100pct'] += 1
    if v['short'] > fix_totals[fix]['peak_short']:
        fix_totals[fix]['peak_short'] = v['short']
        fix_totals[fix]['peak_wk']    = wk

print("{:>8} {:>10} {:>12} {:>10} {:>12} {:>12} {:>8}".format(
    "Fixture","WeekCap","TotalDem","Shortage","Short%","Wks@100%","PeakWk"))
print("-"*75)
for fx in ['123AB','456CD','789EF']:
    v = fix_totals[fx]
    pct = 100*v['short']/v['dem'] if v['dem'] > 0 else 0
    print("{:>8} {:>10,} {:>12,} {:>10,} {:>11.1f}% {:>12} {:>8}".format(
        fx, v['cap'], v['dem'], v['short'], pct, v['weeks_100pct'], v['peak_wk'][:7]))

# 3. Peak shortage weeks per fixture
print("\n[3] TOP-3 SHORTAGE WEEKS BY FIXTURE")
for fix in ['123AB','456CD','789EF']:
    data = [(k[1],v) for k,v in fix_week.items() if k[0]==fix and v['short']>0]
    data.sort(key=lambda x: -x[1]['short'])
    cap = fix_totals[fix]['cap']
    print("  {} (cap={:,}/wk):".format(fix, cap))
    for wk, v in data[:3]:
        gap_pct = 100*v['dem']/cap if cap > 0 else 0
        extra_needed = max(v['dem'] - cap, 0)
        print("    {} Dem={:>6,} Short={:>5,} NeedExtra={:>5,}/wk [{:.0f}% of cap]".format(
            wk, v['dem'], v['short'], extra_needed, gap_pct))

# 4. Upgrade recommendations
print("\n[4] UPGRADE RECOMMENDATIONS")
print("-"*70)

upgrades = [
    {
        "fixture": "456CD",
        "priority": "HIGH",
        "shortage": fix_totals['456CD']['short'],
        "rev_at_risk": (sku_shortage['C']*sku_price.get('C',0) + sku_shortage['D']*sku_price.get('D',0)),
        "current_cap": fix_totals['456CD']['cap'],
        "peak_gap": fix_totals['456CD']['peak_short'],
        "skus": "C (P1), D (P2)",
        "option1": "+1 fixture = +5,000/wk (+100% cap)",
        "option2": "Overtime: +1,000/wk per extra shift",
        "weeks_at_100": fix_totals['456CD']['weeks_100pct'],
    },
    {
        "fixture": "123AB",
        "priority": "MEDIUM",
        "shortage": fix_totals['123AB']['short'],
        "rev_at_risk": (sku_shortage['A']*sku_price.get('A',0) + sku_shortage['B']*sku_price.get('B',0)),
        "current_cap": fix_totals['123AB']['cap'],
        "peak_gap": fix_totals['123AB']['peak_short'],
        "skus": "A (P1), B (P2)",
        "option1": "Optimize scheduling (A/B slot rebalancing)",
        "option2": "+1 fixture = +15,000/wk",
        "weeks_at_100": fix_totals['123AB']['weeks_100pct'],
    },
    {
        "fixture": "789EF",
        "priority": "MEDIUM",
        "shortage": fix_totals['789EF']['short'],
        "rev_at_risk": sku_shortage['F']*sku_price.get('F',0),
        "current_cap": fix_totals['789EF']['cap'],
        "peak_gap": fix_totals['789EF']['peak_short'],
        "skus": "E (P1), F (P2)",
        "option1": "+1 fixture = +3,000/wk (+100% cap)",
        "option2": "Overtime: +500~700/wk",
        "weeks_at_100": fix_totals['789EF']['weeks_100pct'],
    },
]

for u in upgrades:
    print("\n  [{priority}] Fixture {fixture}  |  SKUs: {skus}".format(**u))
    print("    Current cap   : {:,} units/week".format(u['current_cap']))
    print("    Total shortage: {:,} units  |  Revenue at risk: \${:,.0f}".format(u['shortage'], u['rev_at_risk']))
    print("    Weeks at 100% : {}  |  Peak gap: {:,} units/week".format(u['weeks_at_100'], u['peak_gap']))
    print("    Option A      : {}".format(u['option1']))
    print("    Option B      : {}".format(u['option2']))

print()
print("="*70)
print("  SUMMARY: Total shortage = {:,} units | Revenue at risk = \${:,.0f}".format(
    sum(sku_shortage.values()), total_rev))
print("  Priority: 456CD > 789EF > 123AB by shortage volume")
print("="*70)
