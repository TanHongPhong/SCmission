"""
Derive Scenario MPS + MRP from PP output
==========================================
Reads the scenario solver output and creates:
  1. Scenario_MPS: Weekly MPS matching original format
  2. Scenario_MRP: Material explosion for all 15 RMs

Output: adds sheets to SCM_round2.1_scenario_PP_fixed.xlsx
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from collections import defaultdict
import math

RAW_FILE     = "SCM_round2.1_new.xlsx"
SOLVER_FILE  = "SCM_round2.1_scenario_solved.xlsx"
OUTPUT_FILE  = "SCM_round2.1_scenario_PP_fixed.xlsx"

HORIZON_START = date(2026, 4, 15)
HORIZON_END   = date(2026, 5, 15)

# Styles
NAVY    = PatternFill("solid", fgColor="1F4E78")
ORANGE  = PatternFill("solid", fgColor="E65100")
TEAL    = PatternFill("solid", fgColor="00897B")
GREEN   = PatternFill("solid", fgColor="E2F0D9")
YELLOW  = PatternFill("solid", fgColor="FFF2CC")
RED     = PatternFill("solid", fgColor="FDE9D9")
GREY    = PatternFill("solid", fgColor="F2F2F2")
BLUE_L  = PatternFill("solid", fgColor="D6E4F0")
WH_B    = Font(color="FFFFFF", bold=True, size=11)
BLD     = Font(bold=True)
THIN    = Side(style="thin", color="BFBFBF")
BDR     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NFMT    = "#,##0"
PFMT    = "0.0%"


def to_date(x):
    if isinstance(x, datetime): return x.date()
    if isinstance(x, date): return x
    return None


def get_workdays(start, end):
    days = []
    d = start
    while d <= end:
        if d.weekday() != 6:
            days.append(d)
        d += timedelta(days=1)
    return days


def get_week_start(d):
    """Get Monday of the week containing d."""
    return d - timedelta(days=d.weekday())


def main():
    print("=" * 60)
    print("  DERIVING MPS + MRP FROM SCENARIO PP")
    print("=" * 60)

    # ── Read raw data ───────────────────────────────────────
    print("[1/5] Reading data ...")
    raw_wb = load_workbook(RAW_FILE, data_only=True)
    sol_wb = load_workbook(SOLVER_FILE, data_only=True)

    # Read BOM
    bom_df = pd.read_excel(RAW_FILE, sheet_name="BOM & Inventory", header=None)
    bom_data = []
    for r in range(5, 35):
        model = bom_df.iloc[r, 3]
        rm = bom_df.iloc[r, 4]
        if pd.isna(model) or pd.isna(rm): continue
        bom_data.append({
            "sku": str(model).strip(),
            "rm": str(rm).strip(),
            "desc": str(bom_df.iloc[r, 5])[:30] if pd.notna(bom_df.iloc[r, 5]) else "",
            "vendor": str(bom_df.iloc[r, 6])[:10] if pd.notna(bom_df.iloc[r, 6]) else "",
            "bom_usage": float(bom_df.iloc[r, 7]) if pd.notna(bom_df.iloc[r, 7]) else 0,
            "lead_time": int(float(bom_df.iloc[r, 8])) if pd.notna(bom_df.iloc[r, 8]) else 0,
            "stock": int(float(bom_df.iloc[r, 9])) if pd.notna(bom_df.iloc[r, 9]) else 0,
            "pending_po": int(float(bom_df.iloc[r, 10])) if pd.notna(bom_df.iloc[r, 10]) else 0,
            "unit_price": float(bom_df.iloc[r, 11]) if pd.notna(bom_df.iloc[r, 11]) else 0,
            "uom": str(bom_df.iloc[r, 12])[:5] if pd.notna(bom_df.iloc[r, 12]) else "",
        })

    # Read input sheet for SKU info
    inp_df = pd.read_excel(RAW_FILE, sheet_name="input", header=None)
    sku_info = {}
    for r in range(9, 15):
        sku = str(inp_df.iloc[r, 0]).strip()
        sku_info[sku] = {
            "fixture": str(inp_df.iloc[r, 2]).strip(),
            "weekly_cap": int(float(inp_df.iloc[r, 3])),
            "priority": int(float(inp_df.iloc[r, 6])),
        }

    # Read demand matrix for weekly demand
    dem_df = pd.read_excel(RAW_FILE, sheet_name="demand_matrix", header=None)
    week_demand = {}  # {(sku, week_date): qty}
    skus_list = ["A", "B", "C", "D", "E", "F"]
    for r in range(1, dem_df.shape[0]):
        wk = to_date(dem_df.iloc[r, 0])
        if wk is None: continue
        for ci, sku in enumerate(skus_list):
            qty = dem_df.iloc[r, ci + 1]
            if pd.notna(qty):
                week_demand[(sku, wk)] = int(float(qty))

    # Read scenario PP daily production
    sol_pp = sol_wb["Scenario_Result"]
    workdays = get_workdays(HORIZON_START, HORIZON_END)
    num_days = len(workdays)

    sol_rows = []
    for r in range(2, sol_pp.max_row + 1):
        sku = sol_pp.cell(r, 1).value
        if sku is None: break
        rd = {
            "sku": str(sku).strip(),
            "ship": to_date(sol_pp.cell(r, 8).value),
            "demand": int(sol_pp.cell(r, 9).value or 0),
            "daily": [],
        }
        for t in range(num_days):
            rd["daily"].append(int(sol_pp.cell(r, 10 + t).value or 0))
        sc = 10 + num_days
        rd["shortage"] = int(sol_pp.cell(r, sc + 4).value or 0)
        sol_rows.append(rd)

    # Aggregate daily production by SKU per day
    sku_daily = defaultdict(lambda: [0] * num_days)
    for rd in sol_rows:
        for t in range(num_days):
            sku_daily[rd["sku"]][t] += rd["daily"][t]

    # Get weeks in horizon
    all_weeks = sorted(set(get_week_start(d) for d in workdays))
    # Extend to cover full demand matrix weeks
    full_weeks = sorted(set(wk for _, wk in week_demand.keys()))

    # ── Build MPS ───────────────────────────────────────────
    print("[2/5] Building Scenario MPS ...")

    # Aggregate daily production into weekly
    sku_weekly_prod = defaultdict(lambda: defaultdict(int))
    for sku in skus_list:
        for t, d in enumerate(workdays):
            wk = get_week_start(d)
            sku_weekly_prod[sku][wk] += sku_daily[sku][t]

    # Determine scenario weeks (within horizon)
    scenario_weeks = sorted(set(wk for wk in all_weeks if wk >= get_week_start(HORIZON_START)))

    # Open workbook to add sheets
    wb = load_workbook(OUTPUT_FILE)

    # Remove old Scenario_MPS if exists
    if "Scenario_MPS" in wb.sheetnames:
        del wb["Scenario_MPS"]
    if "Scenario_MRP" in wb.sheetnames:
        del wb["Scenario_MRP"]

    ws_mps = wb.create_sheet("Scenario_MPS")

    # MPS headers
    mps_hdrs = ["SKU", "Fixture", "Week", "Weekly Capacity",
                "Total Demand", "Beginning FG Inv", "Storage Cap",
                "Carryover Needed", "Net Required",
                "Planned Production", "Ending FG Inv",
                "Shortage", "ATP", "Capacity Load %", "Planner Note"]

    hdr_fill = ORANGE
    for c, h in enumerate(mps_hdrs, 1):
        cell = ws_mps.cell(1, c, h)
        cell.fill, cell.font = hdr_fill, WH_B
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BDR

    row_idx = 2
    for sku in skus_list:
        info = sku_info[sku]
        weekly_cap = info["weekly_cap"]
        fixture = info["fixture"]
        beg_inv = 0

        for wk in scenario_weeks:
            demand = week_demand.get((sku, wk), 0)
            prod = sku_weekly_prod[sku].get(wk, 0)

            # Carryover: inventory needed for future demand (simplified)
            carryover = max(beg_inv - demand, 0) if beg_inv > demand else 0
            net_req = max(demand - beg_inv, 0)
            available = beg_inv + prod
            end_inv = max(available - demand, 0)
            shortage = max(demand - available, 0)
            atp = end_inv
            load_pct = prod / weekly_cap if weekly_cap > 0 else 0

            # Planner note
            note = ""
            if sku == "A" and wk < date(2026, 5, 4):
                note = "PCBA limited - air only"
            elif sku == "A" and wk >= date(2026, 5, 4):
                note = "PCBA reflash available"
            if shortage > 0:
                note = "SHORTAGE: {:,}".format(shortage)

            vals = [sku, fixture, wk, weekly_cap, demand, beg_inv,
                    info["weekly_cap"], carryover, net_req,
                    prod, end_inv, shortage, atp, load_pct, note]

            for ci, v in enumerate(vals, 1):
                cl = ws_mps.cell(row_idx, ci)
                cl.value = v
                cl.border = BDR
            ws_mps.cell(row_idx, 3).number_format = "DD/MM/YYYY"
            ws_mps.cell(row_idx, 14).number_format = PFMT

            # Color shortage
            if shortage > 0:
                ws_mps.cell(row_idx, 12).fill = RED
                ws_mps.cell(row_idx, 12).font = BLD

            beg_inv = end_inv
            row_idx += 1

    for col_cells in ws_mps.columns:
        cl = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_mps.column_dimensions[cl].width = min(max(mx + 3, 10), 22)
    ws_mps.freeze_panes = "D2"

    # ── Build MRP ───────────────────────────────────────────
    print("[3/5] Building Scenario MRP ...")
    ws_mrp = wb.create_sheet("Scenario_MRP")

    # Get unique RMs
    rm_list = []
    rm_seen = set()
    for b in bom_data:
        if b["rm"] not in rm_seen:
            rm_list.append(b["rm"])
            rm_seen.add(b["rm"])

    # For MRP, use scenario weeks
    mrp_weeks = scenario_weeks

    # Header: Items row
    r = 1
    ws_mrp.cell(r, 1, "Scenario MRP").font = Font(bold=True, size=14)
    ws_mrp.cell(r, 1).fill = NAVY
    ws_mrp.cell(r, 1).font = WH_B

    r = 3
    # Info rows
    info_hdrs = ["Items"] + skus_list + rm_list
    for ci, h in enumerate(info_hdrs, 1):
        cl = ws_mrp.cell(r, ci, h)
        cl.fill, cl.font = ORANGE, WH_B
        cl.border = BDR

    # Lead time row
    r = 4
    ws_mrp.cell(r, 1, "Lead time").font = BLD
    ws_mrp.cell(r, 1).border = BDR
    for ci, sku in enumerate(skus_list, 2):
        ws_mrp.cell(r, ci, "1-2 weeks").border = BDR
    for ci, rm in enumerate(rm_list, 2 + len(skus_list)):
        lt = next((b["lead_time"] for b in bom_data if b["rm"] == rm), 0)
        ws_mrp.cell(r, ci, lt).border = BDR

    # Beginning inventory
    r = 5
    ws_mrp.cell(r, 1, "Beginning inventory").font = BLD
    ws_mrp.cell(r, 1).border = BDR
    for ci, sku in enumerate(skus_list, 2):
        ws_mrp.cell(r, ci, 0).border = BDR
    for ci, rm in enumerate(rm_list, 2 + len(skus_list)):
        stk = next((b["stock"] for b in bom_data if b["rm"] == rm), 0)
        ws_mrp.cell(r, ci, stk).border = BDR
        ws_mrp.cell(r, ci).number_format = NFMT

    # Scheduled receipts
    r = 6
    ws_mrp.cell(r, 1, "Scheduled receipts").font = BLD
    ws_mrp.cell(r, 1).border = BDR
    for ci, sku in enumerate(skus_list, 2):
        ws_mrp.cell(r, ci, "").border = BDR
    for ci, rm in enumerate(rm_list, 2 + len(skus_list)):
        po = next((b["pending_po"] for b in bom_data if b["rm"] == rm), 0)
        ws_mrp.cell(r, ci, po if po > 0 else "").border = BDR

    # Now create MRP for each RM
    r = 8
    for rm in rm_list:
        # Get BOM data for this RM
        rm_bom = [b for b in bom_data if b["rm"] == rm]
        rm_info = rm_bom[0] if rm_bom else {}
        lt = rm_info.get("lead_time", 0)
        lt_weeks = max(1, math.ceil(lt / 7))
        stock = rm_info.get("stock", 0)
        po = rm_info.get("pending_po", 0)
        desc = rm_info.get("desc", "")
        uom = rm_info.get("uom", "")

        # RM04 special: on hold in scenario
        is_pcba = (rm == "RM04")

        # Header
        ws_mrp.cell(r, 1, "Item {}".format(rm)).font = BLD
        ws_mrp.cell(r, 1).fill = TEAL
        ws_mrp.cell(r, 1).font = WH_B
        ws_mrp.cell(r, 2, "LLC: 1").border = BDR
        ws_mrp.cell(r, 3, desc).border = BDR
        r += 1

        # Sub-header
        sub_hdrs = ["", "LT: {} wk".format(lt_weeks)]
        for wk in mrp_weeks:
            sub_hdrs.append(wk)
        for ci, h in enumerate(sub_hdrs, 1):
            cl = ws_mrp.cell(r, ci, h)
            cl.fill, cl.font = BLUE_L, BLD
            cl.border = BDR
            if isinstance(h, date):
                cl.number_format = "DD/MM/YYYY"
        r += 1

        # Gross requirements: sum of (SKU weekly prod * BOM usage)
        gross = {}
        for wk in mrp_weeks:
            total = 0
            for b in rm_bom:
                sku = b["sku"]
                prod = sku_weekly_prod[sku].get(wk, 0)
                total += int(prod * b["bom_usage"])
            gross[wk] = total

        ws_mrp.cell(r, 1, "Gross requirements").border = BDR
        ws_mrp.cell(r, 2, "").border = BDR
        for ci, wk in enumerate(mrp_weeks, 3):
            ws_mrp.cell(r, ci, gross[wk]).border = BDR
            ws_mrp.cell(r, ci).number_format = NFMT
        r += 1

        # Scheduled receipts
        ws_mrp.cell(r, 1, "Scheduled receipts").border = BDR
        ws_mrp.cell(r, 2, "").border = BDR
        for ci, wk in enumerate(mrp_weeks, 3):
            # PO arrives in first week (simplified)
            sr = po if ci == 3 and po > 0 else 0
            if is_pcba:
                sr = 0  # PCBA PO is ON HOLD
            ws_mrp.cell(r, ci, sr if sr > 0 else "").border = BDR
        r += 1

        # Projected on hand
        ws_mrp.cell(r, 1, "Projected on hand").border = BDR
        init_stock = stock if not is_pcba else 0  # PCBA stock ON HOLD
        ws_mrp.cell(r, 2, init_stock).border = BDR
        ws_mrp.cell(r, 2).number_format = NFMT
        oh = init_stock
        oh_vals = {}
        for ci, wk in enumerate(mrp_weeks, 3):
            sr = po if ci == 3 and po > 0 and not is_pcba else 0
            oh = oh + sr - gross[wk]
            oh_vals[wk] = oh
            cl = ws_mrp.cell(r, ci, oh)
            cl.border = BDR
            cl.number_format = NFMT
            if oh < 0:
                cl.fill = RED
                cl.font = BLD
        r += 1

        # Net requirements
        ws_mrp.cell(r, 1, "Net requirements").border = BDR
        ws_mrp.cell(r, 2, "").border = BDR
        net = {}
        cum_deficit = 0
        oh_running = init_stock + (po if not is_pcba else 0)
        for ci, wk in enumerate(mrp_weeks, 3):
            oh_running -= gross[wk]
            nr = max(-oh_running, 0) - cum_deficit
            if nr < 0: nr = 0
            cum_deficit += nr
            oh_running += nr
            net[wk] = nr if nr > 0 else 0
            ws_mrp.cell(r, ci, net[wk] if net[wk] > 0 else "").border = BDR
        r += 1

        # Planned order receipts
        ws_mrp.cell(r, 1, "Planned order receipts").border = BDR
        ws_mrp.cell(r, 2, "").border = BDR
        for ci, wk in enumerate(mrp_weeks, 3):
            ws_mrp.cell(r, ci, net[wk] if net[wk] > 0 else "").border = BDR
        r += 1

        # Planned order releases (offset by lead time)
        ws_mrp.cell(r, 1, "Planned order releases").border = BDR
        ws_mrp.cell(r, 2, "").border = BDR
        for ci, wk in enumerate(mrp_weeks, 3):
            # Offset forward by lt_weeks
            future_idx = ci - 3 + lt_weeks
            if future_idx < len(mrp_weeks):
                future_wk = mrp_weeks[future_idx]
                val = net.get(future_wk, 0)
            else:
                val = 0
            ws_mrp.cell(r, ci, val if val > 0 else "").border = BDR
        r += 1

        # PCBA special note
        if is_pcba:
            ws_mrp.cell(r, 1, "*** RM04 ON HOLD - See PCBA_Recovery_Plan ***").font = Font(bold=True, color="FF0000")
            r += 1

        r += 1  # blank row between RMs

    # Auto width
    for col_cells in ws_mrp.columns:
        cl_letter = get_column_letter(col_cells[0].column)
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws_mrp.column_dimensions[cl_letter].width = min(max(mx + 3, 10), 22)

    # ── Print MPS summary ───────────────────────────────────
    print("[4/5] MPS Summary:")
    for sku in skus_list:
        total_dem = sum(week_demand.get((sku, wk), 0) for wk in scenario_weeks)
        total_prod = sum(sku_weekly_prod[sku].get(wk, 0) for wk in scenario_weeks)
        shortage = max(total_dem - total_prod, 0)
        # Actually compute from carry
        print("    {}: Demand={:>7,}  Prod={:>7,}  Shortage={:>7,}".format(
            sku, total_dem, total_prod, shortage))

    # ── Print MRP summary ───────────────────────────────────
    print("[5/5] MRP Summary (net requirements):")
    for rm in rm_list:
        rm_bom = [b for b in bom_data if b["rm"] == rm]
        total_gross = 0
        for wk in mrp_weeks:
            for b in rm_bom:
                total_gross += int(sku_weekly_prod[b["sku"]].get(wk, 0) * b["bom_usage"])
        stk = next((b["stock"] for b in rm_bom), 0)
        po_val = next((b["pending_po"] for b in rm_bom), 0)
        is_pcba = (rm == "RM04")
        avail = (stk + po_val) if not is_pcba else 0
        net_req = max(total_gross - avail, 0)
        flag = " *** ON HOLD ***" if is_pcba else ""
        print("    {}: Gross={:>8,}  Stock={:>7,}  PO={:>6,}  Net={:>8,}{}".format(
            rm, total_gross, stk, po_val, net_req, flag))

    # ── Save ────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print()
    print("  Sheets added to {}:".format(OUTPUT_FILE))
    print("    - Scenario_MPS ({} rows)".format(row_idx - 2))
    print("    - Scenario_MRP ({} RMs)".format(len(rm_list)))
    print("=" * 60)


if __name__ == "__main__":
    main()
